# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import os
import re
import time
import datetime
from Tkinter import *
import tkMessageBox
import tkFileDialog as filedialog
import ttk
from ttk import *
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import threading
from threading import Thread, Event
from Queue import Queue
import sys
reload(sys)
sys.setdefaultencoding('Cp1252')
import visa
rm = visa.ResourceManager()
rg1 = rm.list_resources()
rg2 = list(rg1)

q = Queue()
sem = threading.Semaphore()
a1 = list()
b1 = list()
ws = 0
wb = 0
inst_1 = 0
inst_2 = 0
data_c2 = 0
thread = 0
protokol = 'C:\ITL\DMM\Protocol'
shablon = 'C:\ITL\DMM\Shablon'


def pribor():
    lb.insert(END, '____________________________________________')
    lb.insert(END, 'Обнаруженные приборы и порты:')
    v = len(rg1)
    i = -1
    while i < v - 1:
        i = i + 1
        if re.search('0A07', rg2[i]):
            rg2[i] = '34411A'
        if re.search('1301', rg2[i]):
            rg2[i] = '34461A'
        if re.search('1401', rg2[i]):
            rg2[i] = '34461A'
        if re.search('0101', rg2[i]):
            rg2[i] = '34465A'
        if re.search('1F01', rg2[i]):
            rg2[i] = 'N5183A'
        if re.search('5707', rg2[i]):
            rg2[i] = '33622A'
        if re.search('5418', rg2[i]):
            rg2[i] = 'N1913A'
        if re.search('0090', rg2[i]):
            rg2[i] = 'CNT-90XL'
        lb.insert(END, rg2[i])
        combo1.configure(values=rg2)
        combo2.configure(values=rg2)
    lb.insert(END, '____________________________________________')


class id:
    def __init__(self, type, adr):
        self.type = type
        self.adr = adr

    def connect_d(self):
        global a1
        global inst_1
        v = len(rg1)
        i = -1
        if combo1.get() == self.type:
            while i < v - 1:
                i = i + 1
                if re.search(self.adr, rg1[i]):
                    inst_1 = rm.open_resource(rg1[i])
                    data_1 = inst_1.query("*IDN?")
                    a.set(data_1)
                    a1 = data_1
                    a1 = re.findall(r'\w+', a1)
                    f.set(a1[2] + ',' + a1[3] + ',' + d + '.xlsx')
                    if a1[2] in ('34410A', '34411A', '34460A', '34461A', '34465A', '34470A'):
                        a10.set('Мультиметр ' + a1[2] + ' ' + 'подключен')
                        lb.insert(END, a10.get())
                        lb.see(END)
                    if a1[2] in ('33622A', 'N5183A'):
                        a10.set('Генератор ' + a1[2] + ' ' + 'подключен')
                        lb.insert(END, a10.get())
                        lb.see(END)

        if combo1.get()[:4] == self.type:
            inst_1 = rm.open_resource(combo1.get())
            inst_1.write(self.adr)
            time.sleep(1)
            data_1 = inst_1.query("*IDN?")
            a.set(data_1)
            a1 = data_1
            a1 = re.findall(r'\w+', a1)
            f.set(a1[2] + ',' + a1[3] + ',' + d + '.xlsx')
            if a1[2] in ('34401A'):
                a10.set('Мультиметр ' + a1[2] + ' ' + 'подключен')
                lb.insert(END, a10.get())
                lb.see(END)


def connect_dmm():
    USB1 = id('34411A', r'\b0x0A07\b').connect_d()
    USB2 = id('34461A', r'\b0x1401\b').connect_d()
    USB3 = id('34461A', r'\b0x1301\b').connect_d()
    USB4 = id('34465A', r'\b0x0101\b').connect_d()
    USB5 = id('33622A', r'\b0x5707\b').connect_d()
    USB6 = id('N1913A', r'\b0x5418\b').connect_d()
    COM = id('ASRL', "SYST:REM").connect_d()


def connect_fluke():
    global b1
    global inst_2
    if combo2.get()[:4] == 'ASRL':
        inst_2 = rm.open_resource(g.get(), baud_rate=9600, data_bits=8, write_termination='\r', read_termination='\r')
        data_2 = inst_2.query("*IDN?")
        b.set(data_2)
        b1 = data_2
        b1 = re.findall(r'\w+', b1)
        b14.set('Калибратор ' + b1[0] + ' ' + b1[1] + ' ' + 'подключен')
        lb.insert(END, b14.get())
        lb.see(END)


class call(Thread):
    def __init__(self, name, d1, volt1, volt2, cell1, cell2, band, time, accurancy):
        Thread.__init__(self)
        self.name = name
        self.d1 = d1
        self.volt1 = volt1
        self.volt2 = volt2
        self.cell1 = cell1
        self.cell2 = cell2
        self.band = band
        self.time = time
        self.accurancy = accurancy
        self.start()

    def run(self):
        sem.acquire()
        inst_1.write(self.volt2)
        inst_1.write(self.band)
        time.sleep(1)
        inst_2.write('*CLS')
        inst_2.write(self.volt1)
        inst_2.write('OPER')
        time.sleep(5)
        inst_1.write('READ?')
        time_1 = float(self.time)
        time.sleep(time_1)
        data_3 = inst_1.read()
        data_4 = float(self.d1)
        data_5 = float(data_3)
        if self.name == 'cap':
            data_c3 = (data_5 - data_c2)
            data_c4 = data_c3 * 10E+8
            data_6 = ((data_c3 - data_4) / data_4) * 100
            ws[self.cell1] = data_c4
        elif self.name == 'res2':
            data_r = data_5 / 10E+5
            data_6 = ((data_5 - data_4) / data_4) * 100
            ws[self.cell1] = data_r
        elif self.name in ('dc', 'ac', 'dci', 'aci', 'fr', 'res4'):
            data_6 = ((data_5 - data_4) / data_4) * 100
            ws[self.cell1] = data_5
        ws[self.cell2] = data_6
        ws['F7'] = a1[2]
        ws['H7'] = a1[3]
        ws['F16'] = h.get()
        ws['F17'] = k.get()
        ws['F18'] = l.get()
        ws['B10'] = m.get()
        data_7 = float(self.accurancy)
        colour = PatternFill(start_color='FFFFDAB9', end_color='FFFFDAB9', fill_type='solid')
        if data_6 > data_7:
            ws[self.cell2].fill = colour
        elif data_6 < -data_7:
            ws[self.cell2].fill = colour
        if a1[2] == '34401A':
            wb.save(protokol + '\\34401A\\' + f.get())
            ws['E97'] = n.get()
            ws['C99'] = e
        elif a1[2] == '34410A':
            wb.save(protokol + '\\34410A\\' + f.get())
            ws['E289'] = n.get()
            ws['C291'] = e
        elif a1[2] == '34411A':
            wb.save(protokol + '\\34411A\\' + f.get())
            ws['E289'] = n.get()
            ws['C291'] = e
        elif a1[2] == '34460A':
            wb.save(protokol + '\\34460A\\' + f.get())
            ws['E121'] = n.get()
            ws['C123'] = e
        elif a1[2] == '34461A':
            wb.save(protokol + '\\34461A\\' + f.get())
            ws['E124'] = n.get()
            ws['C126'] = e
        elif a1[2] == '34465A':
            wb.save(protokol + '\\34465A\\' + f.get())
            ws['E337'] = n.get()
            ws['C339'] = e
        elif a1[2] == '34470A':
            wb.save(protokol + '\\34470A\\' + f.get())
            ws['E337'] = n.get()
            ws['C339'] = e
        inst_2.write('STBY')
        time.sleep(1)
        # progress2.start(60)
        progress1.step(1)
        sem.release()


class reset(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.start()

    def run(self):
        sem.acquire()
        time.sleep(2)
        inst_2.write('*RST')
        inst_1.write('*RST')
        time.sleep(2)
        progress1.step(1)
        sem.release()


class message(Thread):
    def __init__(self, text):
        Thread.__init__(self)
        self.text = text
        self.start()

    def run(self):
        sem.acquire()
        start_thread(q.put(( tkMessageBox.showinfo, ('ВНИМАНИЕ!', self.text), {} )))
        progress1.step(1)
        # progress2.stop()
        sem.release()


class cap(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.start()

    def run(self):
        sem.acquire()
        global data_c2
        inst_1.write('CONF:CAP')
        time.sleep(5)
        inst_1.write('READ?')
        time.sleep(5)
        data_c1 = inst_1.read()
        data_c2 = float(data_c1)
        # inst_1.write('CALC:FUNC NULL')
        # inst_1.write('CALC:STAT ON')
        time.sleep(1)
        progress1.step(1)
        sem.release()


def start():
    global ws
    global wb
    global thread
    enable_animation()
    thread = message('Соедините провода для измерения постоянного напряжения')
    thread = reset()
    if a1[2] in ('34410A', '34411A'):
        wb = load_workbook(shablon + '\\34410A,34411A.xlsx')
        ws = wb.active
        progress1.configure(maximum=221)
        lb.insert(END, 'Время начала:' + ' ' + d1)
        # 0.1V
        thread = call('dc', '0.005', 'OUT 0.005 V', 'CONF:VOLT:DC 0.1', 'C28', 'D28', 'DET:BAND 20', '3', '0.075')
        thread = call('dc', '0.05', 'OUT 0.05 V', 'CONF:VOLT:DC 0.1', 'C29', 'D29', 'DET:BAND 20', '3', '0.012')
        thread = call('dc', '0.095', 'OUT 0.095 V', 'CONF:VOLT:DC 0.1', 'C30', 'D30', 'DET:BAND 20', '3', '0.009')
        # 1V
        thread = call('dc', '0.05', 'OUT 0.05 V', 'CONF:VOLT:DC 1.0', 'C31', 'D31', 'DET:BAND 20', '3', '0.018')
        thread = call('dc', '0.5', 'OUT 0.5 V', 'CONF:VOLT:DC 1.0', 'C32', 'D32', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '0.95', 'OUT 0.95 V', 'CONF:VOLT:DC 1.0', 'C33', 'D33', 'DET:BAND 20', '3', '0.004')
        # 10V
        thread = call('dc', '0.5', 'OUT 0.5 V', 'CONF:VOLT:DC 10', 'C34', 'D34', 'DET:BAND 20', '3', '0.013')
        thread = call('dc', '5', 'OUT 5 V', 'CONF:VOLT:DC 10', 'C35', 'D35', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '9.5', 'OUT 9.5 V', 'CONF:VOLT:DC 10', 'C36', 'D36', 'DET:BAND 20', '3', '0.004')
        # 100V
        thread = call('dc', '5', 'OUT 5 V', 'CONF:VOLT:DC 100', 'C37', 'D37', 'DET:BAND 20', '3', '0.016')
        thread = call('dc', '50', 'OUT 50 V', 'CONF:VOLT:DC 100', 'C38', 'D38', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '95', 'OUT 95 V', 'CONF:VOLT:DC 100', 'C39', 'D39', 'DET:BAND 20', '3', '0.005')
        # 1000V
        thread = call('dc', '50', 'OUT 50 V', 'CONF:VOLT:DC 1000', 'C40', 'D40', 'DET:BAND 20', '3', '0.016')
        thread = call('dc', '500', 'OUT 500 V', 'CONF:VOLT:DC 1000', 'C41', 'D41', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '950', 'OUT 950 V', 'CONF:VOLT:DC 1000', 'C42', 'D42', 'DET:BAND 20', '3', '0.005')	
        thread = reset()
        # ~0.1V,10Hz
        thread = call('ac', '0.005', 'OUT 0.005 V, 10 Hz', 'CONF:VOLT:AC 0.1', 'D47', 'E47', 'DET:BAND 3', '8', '0.7')
        thread = call('ac', '0.05', 'OUT 0.05 V, 10 Hz', 'CONF:VOLT:AC 0.1', 'D48', 'E48', 'DET:BAND 3', '8', '0.16')
        thread = call('ac', '0.095', 'OUT 0.095 V, 10 Hz', 'CONF:VOLT:AC 0.1', 'D49', 'E49', 'DET:BAND 3', '8', '0.132')
        # ~1V,10Hz
        thread = call('ac', '0.05', 'OUT 0.05 V, 10 Hz', 'CONF:VOLT:AC 1.0', 'D50', 'E50', 'DET:BAND 3', '8', '0.7')
        thread = call('ac', '0.5', 'OUT 0.5 V, 10 Hz', 'CONF:VOLT:AC 1.0', 'D51', 'E51', 'DET:BAND 3', '8', '0.16')
        thread = call('ac', '0.95', 'OUT 0.95 V, 10 Hz', 'CONF:VOLT:AC 1.0', 'D52', 'E52', 'DET:BAND 3', '8', '0.132')
        # ~10V,10Hz
        thread = call('ac', '0.5', 'OUT 0.5 V, 10 Hz', 'CONF:VOLT:AC 10.0', 'D53', 'E53', 'DET:BAND 3', '8', '0.7')
        thread = call('ac', '5', 'OUT 5 V, 10 Hz', 'CONF:VOLT:AC 10.0', 'D54', 'E54', 'DET:BAND 3', '8', '0.16')
        thread = call('ac', '9.5', 'OUT 9.5 V, 10 Hz', 'CONF:VOLT:AC 10.0', 'D55', 'E55', 'DET:BAND 3', '8', '0.132')
        # ~100V,10Hz
        thread = call('ac', '5', 'OUT 5 V, 10 Hz', 'CONF:VOLT:AC 100', 'D56', 'E56', 'DET:BAND 3', '8', '0.7')
        # ~0.1V,50Hz
        thread = call('ac', '0.005', 'OUT 0.005 V, 50 Hz', 'CONF:VOLT:AC 0.1', 'D62', 'E62', 'DET:BAND 20', '5', '0.66')
        thread = call('ac', '0.05', 'OUT 0.05 V, 50 Hz', 'CONF:VOLT:AC 0.1', 'D63', 'E63', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '0.095', 'OUT 0.095 V, 50 Hz', 'CONF:VOLT:AC 0.1', 'D64', 'E64', 'DET:BAND 20', '5', '0.092')
        # ~1V,50Hz
        thread = call('ac', '0.05', 'OUT 0.05 V, 50 Hz', 'CONF:VOLT:AC 1.0', 'D65', 'E65', 'DET:BAND 20', '5', '0.66')
        thread = call('ac', '0.5', 'OUT 0.5 V, 50 Hz', 'CONF:VOLT:AC 1.0', 'D66', 'E66', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '0.95', 'OUT 0.95 V, 50 Hz', 'CONF:VOLT:AC 1.0', 'D67', 'E67', 'DET:BAND 20', '5', '0.092')
        # ~10V,50Hz
        thread = call('ac', '0.5', 'OUT 0.5 V, 50 Hz', 'CONF:VOLT:AC 10.0', 'D68', 'E68', 'DET:BAND 20', '5', '0.66')
        thread = call('ac', '5', 'OUT 5 V, 50 Hz', 'CONF:VOLT:AC 10.0', 'D69', 'E69', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '9.5', 'OUT 9.5 V, 50 Hz', 'CONF:VOLT:AC 10.0', 'D70', 'E70', 'DET:BAND 20', '5', '0.092')
        # ~100V,50Hz
        thread = call('ac', '5', 'OUT 5 V, 50 Hz', 'CONF:VOLT:AC 100', 'D71', 'E71', 'DET:BAND 20', '5', '0.66')
        thread = call('ac', '50', 'OUT 50 V, 50 Hz', 'CONF:VOLT:AC 100', 'D72', 'E72', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '95', 'OUT 95 V, 50 Hz', 'CONF:VOLT:AC 100', 'D73', 'E73', 'DET:BAND 20', '5', '0.092')	
        # ~750V,50Hz
        thread = call('ac', '37.5', 'OUT 37.5 V, 50 Hz', 'CONF:VOLT:AC 750.0', 'D74', 'E74', 'DET:BAND 20', '5', '0.66')
        thread = call('ac', '375', 'OUT 375 V, 50 Hz', 'CONF:VOLT:AC 750.0', 'D75', 'E75', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '712.5', 'OUT 712.5 V, 50 Hz', 'CONF:VOLT:AC 750.0', 'D76', 'E76', 'DET:BAND 20', '5', '0.092')	
        # ~0.1V,20кHz
        thread = call('ac', '0.005', 'OUT 0.005 V, 20000 Hz', 'CONF:VOLT:AC 0.1', 'D77', 'E77', 'DET:BAND 20', '5', '1.1')
        thread = call('ac', '0.05', 'OUT 0.05 V, 20000 Hz', 'CONF:VOLT:AC 0.1', 'D78', 'E78', 'DET:BAND 20', '5', '0.2')
        thread = call('ac', '0.095', 'OUT 0.095 V, 20000 Hz', 'CONF:VOLT:AC 0.1', 'D79', 'E79', 'DET:BAND 20', '5', '0.153')
        # ~1V,20kHz
        thread = call('ac', '0.05', 'OUT 0.05 V, 20000 Hz', 'CONF:VOLT:AC 1.0', 'D80', 'E80', 'DET:BAND 20', '5', '1.1')
        thread = call('ac', '0.5', 'OUT 0.5 V, 20000 Hz', 'CONF:VOLT:AC 1.0', 'D81', 'E81', 'DET:BAND 20', '5', '0.2')
        thread = call('ac', '0.95', 'OUT 0.95 V, 20000 Hz', 'CONF:VOLT:AC 1.0', 'D82', 'E82', 'DET:BAND 20', '5', '0.153')
        # ~10V,20kHz
        thread = call('ac', '0.5', 'OUT 0.5 V, 20000 Hz', 'CONF:VOLT:AC 10.0', 'D83', 'E83', 'DET:BAND 20', '5', '1.1')
        thread = call('ac', '5', 'OUT 5 V, 20000 Hz', 'CONF:VOLT:AC 10.0', 'D84', 'E84', 'DET:BAND 20', '5', '0.2')
        thread = call('ac', '9.5', 'OUT 9.5 V, 20000 Hz', 'CONF:VOLT:AC 10.0', 'D85', 'E85', 'DET:BAND 20', '5', '0.153')
        # ~100V,20kHz
        thread = call('ac', '5', 'OUT 5 V, 20000 Hz', 'CONF:VOLT:AC 100', 'D86', 'E86', 'DET:BAND 20', '5', '1.1')
        thread = call('ac', '50', 'OUT 50 V, 20000 Hz', 'CONF:VOLT:AC 100', 'D87', 'E87', 'DET:BAND 200', '5', '0.2')
        thread = call('ac', '95', 'OUT 95 V, 20000 Hz', 'CONF:VOLT:AC 100', 'D88', 'E88', 'DET:BAND 20', '5', '0.153')
        # ~750V,20kHz
        thread = call('ac', '37.5', 'OUT 37.5 V, 20000 Hz', 'CONF:VOLT:AC 750', 'D89', 'E89', 'DET:BAND 20', '5', '1.1')
        # ~0.1V,50кHz
        thread = call('ac', '0.005', 'OUT 0.005 V, 50000 Hz', 'CONF:VOLT:AC 0.1', 'D92', 'E92', 'DET:BAND 20', '5', '2.0')
        thread = call('ac', '0.05', 'OUT 0.05 V, 50000 Hz', 'CONF:VOLT:AC 0.1', 'D93', 'E93', 'DET:BAND 20', '5', '0.56')
        thread = call('ac', '0.095', 'OUT 0.095 V, 50000 Hz', 'CONF:VOLT:AC 0.1', 'D94', 'E94', 'DET:BAND 20', '5', '0.484')
        # ~1V,50kHz
        thread = call('ac', '0.05', 'OUT 0.05 V, 50000 Hz', 'CONF:VOLT:AC 1.0', 'D95', 'E95', 'DET:BAND 20', '5', '2.0')
        thread = call('ac', '0.5', 'OUT 0.5 V, 50000 Hz', 'CONF:VOLT:AC 1.0', 'D96', 'E96', 'DET:BAND 20', '5', '0.56')
        thread = call('ac', '0.95', 'OUT 0.95 V, 50000 Hz', 'CONF:VOLT:AC 1.0', 'D97', 'E97', 'DET:BAND 20', '5', '0.484')
        # ~10V,50kHz
        thread = call('ac', '0.5', 'OUT 0.5 V, 50000 Hz', 'CONF:VOLT:AC 10.0', 'D98', 'E98', 'DET:BAND 20', '5', '2.0')
        thread = call('ac', '5', 'OUT 5 V, 50000 Hz', 'CONF:VOLT:AC 10.0', 'D99', 'E99', 'DET:BAND 20', '5', '0.56')
        thread = call('ac', '9.5', 'OUT 9.5 V, 50000 Hz', 'CONF:VOLT:AC 10.0', 'D100', 'E100', 'DET:BAND 20', '5', '0.484')
        # ~100V,50kHz
        thread = call('ac', '5', 'OUT 5 V, 50000 Hz', 'CONF:VOLT:AC 100', 'D101', 'E101', 'DET:BAND 20', '5', '2.0')
        # ~0.1V,100кHz
        thread = call('ac', '0.005', 'OUT 0.005 V, 100000 Hz', 'CONF:VOLT:AC 0.1', 'D107', 'E107', 'DET:BAND 20', '5', '11.2')
        thread = call('ac', '0.05', 'OUT 0.05 V, 100000 Hz', 'CONF:VOLT:AC 0.1', 'D108', 'E108', 'DET:BAND 20', '5', '2.2')
        thread = call('ac', '0.095', 'OUT 0.095 V, 100000 Hz', 'CONF:VOLT:AC 0.1', 'D109', 'E109', 'DET:BAND 20', '5', '1.726')
        # ~1V,100kHz
        thread = call('ac', '0.05', 'OUT 0.05 V, 100000 Hz', 'CONF:VOLT:AC 1.0', 'D110', 'E110', 'DET:BAND 20', '5', '11.2')
        thread = call('ac', '0.5', 'OUT 0.5 V, 100000 Hz', 'CONF:VOLT:AC 1.0', 'D111', 'E111', 'DET:BAND 20', '5', '2.2')
        thread = call('ac', '0.95', 'OUT 0.95 V, 100000 Hz', 'CONF:VOLT:AC 1.0', 'D112', 'E112', 'DET:BAND 20', '5', '1.726')
        # ~10V,100kHz
        thread = call('ac', '0.5', 'OUT 0.5 V, 100000 Hz', 'CONF:VOLT:AC 10.0', 'D113', 'E113', 'DET:BAND 20', '5', '11.2')
        thread = call('ac', '5', 'OUT 5 V, 100000 Hz', 'CONF:VOLT:AC 10.0', 'D114', 'E114', 'DET:BAND 20', '5', '2.2')
        thread = call('ac', '9.5', 'OUT 9.5 V, 100000 Hz', 'CONF:VOLT:AC 10.0', 'D115', 'E115', 'DET:BAND 20', '5', '1.726')
        # ~100V,100kHz
        thread = call('ac', '5', 'OUT 5 V, 100000 Hz', 'CONF:VOLT:AC 100', 'D116', 'E116', 'DET:BAND 20', '5', '11.2')
        # ~0.1V,300kHz
        thread = call('ac', '0.005', 'OUT 0.005 V, 300000 Hz', 'CONF:VOLT:AC 0.1', 'D122', 'E122', 'DET:BAND 3', '5', '11.2')
        thread = call('ac', '0.05', 'OUT 0.05 V, 300000 Hz', 'CONF:VOLT:AC 0.1', 'D123', 'E123', 'DET:BAND 3', '5', '2.2')
        thread = call('ac', '0.095', 'OUT 0.095 V, 300000 Hz', 'CONF:VOLT:AC 0.1', 'D124', 'E124', 'DET:BAND 3', '5', '1.726')
        # ~1V,300kHz
        thread = call('ac', '0.05', 'OUT 0.05 V, 300000 Hz', 'CONF:VOLT:AC 1.0', 'D125', 'E125', 'DET:BAND 20', '5', '11.2')
        thread = call('ac', '0.5', 'OUT 0.5 V, 300000 Hz', 'CONF:VOLT:AC 1.0', 'D126', 'E126', 'DET:BAND 20', '5', '2.2')
        thread = call('ac', '0.95', 'OUT 0.95 V, 300000 Hz', 'CONF:VOLT:AC 1.0', 'D127', 'E127', 'DET:BAND 20', '5', '1.726')
        # ~10V,300kHz
        thread = call('ac', '0.5', 'OUT 0.5 V, 300000 Hz', 'CONF:VOLT:AC 10', 'D128', 'E128', 'DET:BAND 20', '5', '11.2')
        thread = reset()
        # 5Hz
        thread = call('fr', '5.0', 'OUT 0.1 V, 5.0 Hz', 'CONF:FREQ 5.0 Hz', 'C222', 'D222', 'DET:BAND 20', '5', '0.07')
        thread = call('fr', '5.0', 'OUT 10.0 V, 5.0 Hz', 'CONF:FREQ 5.0 Hz', 'C223', 'D223', 'DET:BAND 20', '5', '0.07')
        # 10Hz
        thread = call('fr', '10.0', 'OUT 0.1 V, 10.0 Hz', 'CONF:FREQ 10.0 Hz', 'C224', 'D224', 'DET:BAND 20', '5', '0.04')
        thread = call('fr', '10.0', 'OUT 10.0 V, 10.0 Hz', 'CONF:FREQ 10.0 Hz', 'C225', 'D225', 'DET:BAND 20', '5', '0.04')
        # 40Hz
        thread = call('fr', '40.0', 'OUT 0.1 V, 40.0 Hz', 'CONF:FREQ 40.0 Hz', 'C226', 'D226', 'DET:BAND 20', '5', '0.02')
        thread = call('fr', '40.0', 'OUT 10.0 V, 40.0 Hz', 'CONF:FREQ 40.0 Hz', 'C227', 'D227', 'DET:BAND 20', '5', '0.02')	
        # 1kHz
        thread = call('fr', '1000.0', 'OUT 0.1 V, 1.0 kHz', 'CONF:FREQ 1.0 kHz', 'C228', 'D228', 'DET:BAND 20', '5', '0.005')
        thread = call('fr', '1000.0', 'OUT 10.0 V, 1.0 kHz', 'CONF:FREQ 1.0 kHz', 'C229', 'D229', 'DET:BAND 20', '5', '0.005')
        # 100kHz
        thread = call('fr', '100000.0', 'OUT 0.1 V, 100.0 kHz', 'CONF:FREQ 100.0 kHz', 'C230', 'D230', 'DET:BAND 20', '5', '0.005')
        thread = call('fr', '100000.0', 'OUT 10.0 V, 100.0 kHz', 'CONF:FREQ 100.0 kHz', 'C231', 'D231', 'DET:BAND 20', '5', '0.005')
        # 300kHz
        thread = call('fr', '300000.0', 'OUT 0.1 V, 300.0 kHz', 'CONF:FREQ 300.0 kHz', 'C232', 'D232', 'DET:BAND 20', '5', '0.005')
        thread = call('fr', '300000.0', 'OUT 10.0 V, 300.0 kHz', 'CONF:FREQ 300.0 kHz', 'C233', 'D233', 'DET:BAND 20', '5', '0.005')
        # 1NF
        thread = message('Измерение ёмкости.\nВытащите красный провод из каллибратора\nдля компенсации проводов')
        thread = cap()
        thread = message('Верните провод на место')
        thread = call('cap', '0.35E-9', 'OUT 0.35 NF', 'CONF:CAP 1 NF', 'C272', 'D272', 'DET:BAND 20', '5', '1.929')
        thread = call('cap', '0.5E-9', 'OUT 0.5 NF', 'CONF:CAP 1 NF', 'C273', 'D273', 'DET:BAND 20', '5', '1.5')
        thread = call('cap', '0.95E-9', 'OUT 0.95 NF', 'CONF:CAP 1 NF', 'C274', 'D274', 'DET:BAND 20', '5', '1.026')
        # 10NF
        thread = call('cap', '0.5E-9', 'OUT 0.5 NF', 'CONF:CAP 10 NF', 'C275', 'D275', 'DET:BAND 20', '5', '2.4')
        thread = call('cap', '5.0E-9', 'OUT 5.0 NF', 'CONF:CAP 10 NF', 'C276', 'D276', 'DET:BAND 20', '5', '0.6')
        thread = call('cap', '9.5E-9', 'OUT 9.5 NF', 'CONF:CAP 10 NF', 'C277', 'D277', 'DET:BAND 20', '5', '0.505')
        # 100NF
        thread = call('cap', '5.0E-9', 'OUT 5.0 NF', 'CONF:CAP 100 NF', 'C278', 'D278', 'DET:BAND 20', '5', '2.4')
        thread = call('cap', '50.0E-9', 'OUT 50.0 NF', 'CONF:CAP 100 NF', 'C279', 'D279', 'DET:BAND 20', '5', '0.6')
        thread = call('cap', '95.0E-9', 'OUT 95.0 NF', 'CONF:CAP 100 NF', 'C280', 'D280', 'DET:BAND 20', '5', '0.505')
        # 1000NF
        thread = call('cap', '50.0E-9', 'OUT 50.0 NF', 'CONF:CAP 1 UF', 'C281', 'D281', 'DET:BAND 20', '5', '2.4')
        thread = call('cap', '500.0E-9', 'OUT 500.0 NF', 'CONF:CAP 1 UF', 'C282', 'D282', 'DET:BAND 20', '5', '0.6')
        thread = call('cap', '950.0E-9', 'OUT 950.0 NF', 'CONF:CAP 1 UF', 'C283', 'D283', 'DET:BAND 20', '5', '0.505')
        # 10000NF
        thread = call('cap', '500.0E-9', 'OUT 500.0 NF', 'CONF:CAP 10 UF', 'C284', 'D284', 'DET:BAND 20', '5', '2.4')
        thread = call('cap', '5000.0E-9', 'OUT 5000.0 NF', 'CONF:CAP 10 UF', 'C285', 'D285', 'DET:BAND 20', '5', '0.6')
        thread = call('cap', '9500.0E-9', 'OUT 9500.0 NF', 'CONF:CAP 10 UF', 'C286', 'D286', 'DET:BAND 20', '5', '0.505')	
        thread = message('Переключите провода\n для измерения тока')
        thread = reset()
        # 0.0001A
        thread = call('dci', '0.000005', 'OUT 0.000005 A', 'CONF:CURR:DC 0.0001', 'C141', 'D141', 'DET:BAND 20', '5', '0.55')
        thread = call('dci', '0.00005', 'OUT 0.00005 A', 'CONF:CURR:DC 0.0001', 'C142', 'D142', 'DET:BAND 20', '5', '0.1')
        thread = call('dci', '0.000095', 'OUT 0.000095 A', 'CONF:CURR:DC 0.0001', 'C143', 'D143', 'DET:BAND 20', '5', '0.076')
        # 0.001A
        thread = call('dci', '0.00005', 'OUT 0.00005 A', 'CONF:CURR:DC 0.001', 'C144', 'D144', 'DET:BAND 20', '5', '0.17')
        thread = call('dci', '0.0005', 'OUT 0.0005 A', 'CONF:CURR:DC 0.001', 'C145', 'D145', 'DET:BAND 20', '5', '0.062')
        thread = call('dci', '0.00095', 'OUT 0.00095 A', 'CONF:CURR:DC 0.001', 'C146', 'D146', 'DET:BAND 20', '5', '0.056')
        # 0.01A
        thread = call('dci', '0.0005', 'OUT 0.0005 A', 'CONF:CURR:DC 0.01', 'C147', 'D147', 'DET:BAND 20', '5', '0.45')
        thread = call('dci', '0.005', 'OUT 0.005 A', 'CONF:CURR:DC 0.01', 'C148', 'D148', 'DET:BAND 20', '5', '0.09')
        thread = call('dci', '0.0095', 'OUT 0.0095 A', 'CONF:CURR:DC 0.01', 'C149', 'D149', 'DET:BAND 20', '5', '0.071')
        # 0.1A
        thread = call('dci', '0.005', 'OUT 0.005 A', 'CONF:CURR:DC 0.1', 'C150', 'D150', 'DET:BAND 20', '5', '0.15')
        thread = call('dci', '0.05', 'OUT 0.05 A', 'CONF:CURR:DC 0.1', 'C151', 'D151', 'DET:BAND 20', '5', '0.06')
        thread = call('dci', '0.095', 'OUT 0.095 A', 'CONF:CURR:DC 0.1', 'C152', 'D152', 'DET:BAND 20', '5', '0.055')
        # 1A
        thread = call('dci', '0.05', 'OUT 0.05 A', 'CONF:CURR:DC 1.0', 'C153', 'D153', 'DET:BAND 20', '5', '0.3')
        thread = call('dci', '0.5', 'OUT 0.5 A', 'CONF:CURR:DC 1.0', 'C154', 'D154', 'DET:BAND 20', '5', '0.12')
        thread = call('dci', '0.95', 'OUT 0.95 A', 'CONF:CURR:DC 1.0', 'C155', 'D155', 'DET:BAND 20', '5', '0.111')
        # 3A
        thread = call('dci', '0.15', 'OUT 0.15 A', 'CONF:CURR:DC 3.0', 'C156', 'D156', 'DET:BAND 20', '5', '0.55')
        thread = call('dci', '1.5', 'OUT 1.5 A', 'CONF:CURR:DC 3.0', 'C157', 'D157', 'DET:BAND 20', '5', '0.19')
        if b1[1] == '5500E':
            thread = call('dci', '2.85', 'OUT 2.85 A', 'CONF:CURR:DC 3.0', 'C158', 'D158', 'DET:BAND 20', '5', '0.171')
        thread = reset()
        # ~0.0001A, 50Hz
        thread = call('aci', '0.00003', 'OUT 0.00003 A, 50 Hz', 'CONF:CURR:AC 0.0001', 'D163', 'E163', 'DET:BAND 20', '5', '0.143')
        thread = call('aci', '0.00005', 'OUT 0.00005 A, 50 Hz', 'CONF:CURR:AC 0.0001', 'D164', 'E164', 'DET:BAND 20', '5', '0.09')
        thread = call('aci', '0.000095', 'OUT 0.000095 A, 50 Hz', 'CONF:CURR:AC 0.0001', 'D165', 'E165', 'DET:BAND 20', '5', '0.052')
        # ~0.001A, 50Hz
        thread = call('aci', '0.00005', 'OUT 0.00005 A, 50 Hz', 'CONF:CURR:AC 0.001', 'D166', 'E166', 'DET:BAND 20', '5', '0.81')
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 50 Hz', 'CONF:CURR:AC 0.001', 'D167', 'E167', 'DET:BAND 20', '5', '0.09')
        thread = call('aci', '0.00095', 'OUT 0.00095 A, 50 Hz', 'CONF:CURR:AC 0.001', 'D168', 'E168', 'DET:BAND 20', '5', '0.052')
        # ~0.01A, 50Hz
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 50 Hz', 'CONF:CURR:AC 0.01', 'D169', 'E169', 'DET:BAND 20', '5', '0.81')
        thread = call('aci', '0.005', 'OUT 0.005 A, 50 Hz', 'CONF:CURR:AC 0.01', 'D170', 'E170', 'DET:BAND 20', '5', '0.09')
        thread = call('aci', '0.0095', 'OUT 0.0095 A, 50 Hz', 'CONF:CURR:AC 0.01', 'D171', 'E171', 'DET:BAND 20', '5', '0.052')
        # ~0.1A, 50Hz
        thread = call('aci', '0.005', 'OUT 0.005 A, 50 Hz', 'CONF:CURR:AC 0.1', 'D172', 'E172', 'DET:BAND 20', '5', '0.81')
        thread = call('aci', '0.05', 'OUT 0.05 A, 50 Hz', 'CONF:CURR:AC 0.1', 'D173', 'E173', 'DET:BAND 20', '5', '0.09')
        thread = call('aci', '0.095', 'OUT 0.095 A, 50 Hz', 'CONF:CURR:AC 0.1', 'D174', 'E174', 'DET:BAND 20', '5', '0.052')
        # ~1A, 50Hz
        thread = call('aci', '0.05', 'OUT 0.05 A, 50 Hz', 'CONF:CURR:AC 1.0', 'D175', 'E175', 'DET:BAND 20', '5', '0.81')
        thread = call('aci', '0.5', 'OUT 0.5 A, 50 Hz', 'CONF:CURR:AC 1.0', 'D176', 'E176', 'DET:BAND 20', '5', '0.09')
        thread = call('aci', '0.95', 'OUT 0.95 A, 50 Hz', 'CONF:CURR:AC 1.0', 'D177', 'E177', 'DET:BAND 20', '5', '0.052')
        # ~3A, 50Hz
        thread = call('aci', '0.15', 'OUT 0.15 A, 50 Hz', 'CONF:CURR:AC 3.0', 'D178', 'E178', 'DET:BAND 20', '5', '0.81')
        thread = call('aci', '1.5', 'OUT 1.5 A, 50 Hz', 'CONF:CURR:AC 3.0', 'D179', 'E179', 'DET:BAND 20', '5', '0.09')
        if b1[1] == '5500E':
            thread = call('aci', '2.85', 'OUT 2.85 A, 50 Hz', 'CONF:CURR:AC 3.0', 'D180', 'E180', 'DET:BAND 20', '8', '0.052')
        # ~0.0001A, 5kHz
        thread = call('aci', '0.00003', 'OUT 0.00003 A, 5 kHz', 'CONF:CURR:AC 0.0001', 'D181', 'E181', 'DET:BAND 20', '5', '0.153')
        thread = call('aci', '0.00005', 'OUT 0.00005 A, 5 kHz', 'CONF:CURR:AC 0.0001', 'D182', 'E182', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.000095', 'OUT 0.000095 A, 5 kHz', 'CONF:CURR:AC 0.0001', 'D183', 'E183', 'DET:BAND 20', '5', '0.062')
        # ~0.001A, 5kHz
        thread = call('aci', '0.00005', 'OUT 0.00005 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D184', 'E184', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D185', 'E185', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.00095', 'OUT 0.00095 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D186', 'E186', 'DET:BAND 20', '5', '0.062')
        # ~0.01A, 5kHz
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D187', 'E187', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '0.005', 'OUT 0.005 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D188', 'E188', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.0095', 'OUT 0.0095 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D189', 'E189', 'DET:BAND 20', '5', '0.062')
        # ~0.1A, 5kHz
        thread = call('aci', '0.005', 'OUT 0.005 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D190', 'E190', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '0.05', 'OUT 0.05 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D191', 'E191', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.095', 'OUT 0.095 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D192', 'E192', 'DET:BAND 20', '5', '0.062')
        # ~1A, 5kHz
        thread = call('aci', '0.05', 'OUT 0.05 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D193', 'E193', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '0.5', 'OUT 0.5 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D194', 'E194', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.95', 'OUT 0.95 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D195', 'E195', 'DET:BAND 20', '5', '0.062')
        # ~3A, 5kHz
        thread = call('aci', '0.15', 'OUT 0.15 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D196', 'E196', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '1.5', 'OUT 1.5 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D197', 'E197', 'DET:BAND 20', '5', '0.1')
        if b1[1] == '5500E':
            thread = call('aci', '2.85', 'OUT 2.85 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D198', 'E198', 'DET:BAND 20', '5', '0.062')
        # ~0.0001A, 10kHz
        thread = call('aci', '0.00003', 'OUT 0.00003 A, 10 kHz', 'CONF:CURR:AC 0.0001', 'D199', 'E199', 'DET:BAND 20', '5', '0.153')
        thread = call('aci', '0.00005', 'OUT 0.00005 A, 10 kHz', 'CONF:CURR:AC 0.0001', 'D200', 'E200', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.000095', 'OUT 0.000095 A, 10 kHz', 'CONF:CURR:AC 0.0001', 'D201', 'E201', 'DET:BAND 20', '5', '0.062')
        # ~0.001A, 10kHz
        thread = call('aci', '0.00005', 'OUT 0.00005 A, 10 kHz', 'CONF:CURR:AC 0.001', 'D202', 'E202', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 10 kHz', 'CONF:CURR:AC 0.001', 'D203', 'E203', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.00095', 'OUT 0.00095 A, 10 kHz', 'CONF:CURR:AC 0.001', 'D204', 'E204', 'DET:BAND 20', '5', '0.062')
        # ~0.01A, 10kHz
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 10 kHz', 'CONF:CURR:AC 0.01', 'D205', 'E205', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '0.005', 'OUT 0.005 A, 10 kHz', 'CONF:CURR:AC 0.01', 'D206', 'E206', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.0095', 'OUT 0.0095 A, 10 kHz', 'CONF:CURR:AC 0.01', 'D207', 'E207', 'DET:BAND 20', '5', '0.062')
        # ~0.1A, 10kHz
        thread = call('aci', '0.005', 'OUT 0.005 A, 10 kHz', 'CONF:CURR:AC 0.1', 'D208', 'E208', 'DET:BAND 20', '5', '0.82')
        thread = call('aci', '0.05', 'OUT 0.05 A, 10 kHz', 'CONF:CURR:AC 0.1', 'D209', 'E209', 'DET:BAND 20', '5', '0.1')
        thread = call('aci', '0.095', 'OUT 0.095 A, 10 kHz', 'CONF:CURR:AC 0.1', 'D210', 'E210', 'DET:BAND 20', '5', '0.062')
        # ~1A, 10kHz
        thread = call('aci', '0.05', 'OUT 0.05 A, 10 kHz', 'CONF:CURR:AC 1.0', 'D211', 'E211', 'DET:BAND 20', '5', '0.82')
        if b1[1] == '5500E':
            thread = call('aci', '0.5', 'OUT 0.5 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D212', 'E212', 'DET:BAND 20', '5', '0.1')
            thread = call('aci', '0.95', 'OUT 0.95 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D213', 'E213', 'DET:BAND 20', '5', '0.062')
        elif b1[1] == '5522A':
            thread = call('aci', '0.5', 'OUT 0.5 A, 10 kHz', 'CONF:CURR:AC 1.0', 'D212', 'E212', 'DET:BAND 20', '5', '0.1')
            thread = call('aci', '0.95', 'OUT 0.95 A, 10 kHz', 'CONF:CURR:AC 1.0', 'D213', 'E213', 'DET:BAND 20', '5', '0.062')
        # ~3A, 10kHz
        thread = call('aci', '0.15', 'OUT 0.15 A, 10 kHz', 'CONF:CURR:AC 3.0', 'D214', 'E214', 'DET:BAND 20', '5', '0.82')
        if b1[1] == '5500E':
            thread = call('aci', '1.5', 'OUT 1.5 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D215', 'E215', 'DET:BAND 20', '5', '0.1')
            thread = call('aci', '2.85', 'OUT 2.85 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D216', 'E216', 'DET:BAND 20', '5', '0.062')
        elif b1[1] == '5522A':
            thread = call('aci', '1.5', 'OUT 1.5 A, 10 kHz', 'CONF:CURR:AC 3.0', 'D215', 'E215', 'DET:BAND 20', '5', '0.1')
        if b1[1] == '5522A':
            thread = message('Переключите красный провод на калибраторе в разъем больше 2,5 А')
            thread = call('dci', '2.85', 'OUT 2.85 A', 'CONF:CURR:DC 3.0', 'C158', 'D158', 'DET:BAND 20', '5', '0.171')
            thread = call('aci', '2.85', 'OUT 2.85 A, 50 Hz', 'CONF:CURR:AC 3.0', 'D180', 'E180', 'DET:BAND 20', '8', '0.052')
            thread = call('aci', '2.85', 'OUT 2.85 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D198', 'E198', 'DET:BAND 20', '5', '0.062')
            thread = call('aci', '2.85', 'OUT 2.85 A, 10 kHz', 'CONF:CURR:AC 3.0', 'D216', 'E216', 'DET:BAND 20', '5', '0.062')
        thread = message('Переключите провода по четырехпроводной схеме\n для измерения сопротивления')
        thread = reset()
        # 100Ohm-4-wire
        thread = call('res4', '5', 'OUT 5 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'C238', 'D238', 'DET:BAND 20', '5', '0.09')
        thread = call('res4', '50', 'OUT 50 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'C239', 'D239', 'DET:BAND 20', '5', '0.018')
        thread = call('res4', '95', 'OUT 95 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'C240', 'D240', 'DET:BAND 20', '5', '0.014')
        # 1kOhm-4-wire
        thread = call('res4', '50', 'OUT 50 OHM; ZCOMP WIRE4', 'CONF:FRES 1 KOHM', 'C241', 'D241', 'DET:BAND 20', '5', '0.03')
        thread = call('res4', '500', 'OUT 500 OHM; ZCOMP WIRE4', 'CONF:FRES 1 KOHM', 'C242', 'D242', 'DET:BAND 20', '5', '0.012')
        thread = call('res4', '950', 'OUT 950 OHM; ZCOMP WIRE4', 'CONF:FRES 1 KOHM', 'C243', 'D243', 'DET:BAND 20', '5', '0.011')
        # 10kOhm-4-wire
        thread = call('res4', '500', 'OUT 500 OHM; ZCOMP WIRE4', 'CONF:FRES 10 KOHM', 'C244', 'D244', 'DET:BAND 20', '5', '0.03')
        thread = call('res4', '5000', 'OUT 5000 OHM; ZCOMP WIRE4', 'CONF:FRES 10 KOHM', 'C245', 'D245', 'DET:BAND 20', '5', '0.012')
        thread = call('res4', '9500', 'OUT 9500 OHM; ZCOMP WIRE4', 'CONF:FRES 10 KOHM', 'C246', 'D246', 'DET:BAND 20', '5', '0.011')
        # 100kOhm-4-wire
        thread = call('res4', '5000', 'OUT 5000 OHM; ZCOMP WIRE4', 'CONF:FRES 100 KOHM', 'C247', 'D247', 'DET:BAND 20', '5', '0.03')
        thread = call('res4', '50000', 'OUT 50000 OHM; ZCOMP WIRE4', 'CONF:FRES 100 KOHM', 'C248', 'D248', 'DET:BAND 20', '5', '0.012')
        thread = call('res4', '95000', 'OUT 95000 OHM; ZCOMP WIRE4', 'CONF:FRES 100 KOHM', 'C249', 'D249', 'DET:BAND 20', '5', '0.011')
        thread = message('Переключите провода по двухпроводной схеме\n для измерения сопротивления')
        thread = reset()
        # 1MOhm-2-wire
        thread = call('res2', '50000', 'OUT 0.05 MOHM; ZCOMP WIRE2', 'CONF:RES 1 MOHM', 'C255', 'D255', 'DET:BAND 20', '5', '0.03')
        thread = call('res2', '500000', 'OUT 0.5 MOHM; ZCOMP WIRE2', 'CONF:RES 1 MOHM', 'C256', 'D256', 'DET:BAND 20', '5', '0.012')
        thread = call('res2', '950000', 'OUT 0.95 MOHM; ZCOMP WIRE2', 'CONF:RES 1 MOHM', 'C257', 'D257', 'DET:BAND 20', '5', '0.011')
        # 10MOhm-2-wire
        thread = call('res2', '500000', 'OUT 0.5 MOHM; ZCOMP WIRE2', 'CONF:RES 10 MOHM', 'C258', 'D258', 'DET:BAND 20', '5', '0.03')
        thread = call('res2', '5000000', 'OUT 5.0 MOHM; ZCOMP WIRE2', 'CONF:RES 10 MOHM', 'C259', 'D259', 'DET:BAND 20', '5', '0.012')
        thread = call('res2', '9500000', 'OUT 9.5 MOHM; ZCOMP WIRE2', 'CONF:RES 10 MOHM', 'C260', 'D260', 'DET:BAND 20', '5', '0.011')
        # 100MOhm-2-wire
        thread = call('res2', '5000000', 'OUT 5.0 MOHM; ZCOMP WIRE2', 'CONF:RES 100 MOHM', 'C261', 'D261', 'DET:BAND 20', '5', '0.03')
        thread = call('res2', '50000000', 'OUT 50.0 MOHM; ZCOMP WIRE2', 'CONF:RES 100 MOHM', 'C262', 'D262', 'DET:BAND 20', '5', '0.012')
        thread = call('res2', '95000000', 'OUT 95.0 MOHM; ZCOMP WIRE2', 'CONF:RES 100 MOHM', 'C263', 'D263', 'DET:BAND 20', '5', '0.011')
        # 1GOhm-2-wire
        thread = call('res2', '50000000', 'OUT 50.0 MOHM; ZCOMP WIRE2', 'CONF:RES 1 GOHM', 'C264', 'D264', 'DET:BAND 20', '5', '0.03')
        if b1[1] == '5522A':
            thread = call('res2', '500000000', 'OUT 500.0 MOHM; ZCOMP WIRE2', 'CONF:RES 1 GOHM', 'C265', 'D265', 'DET:BAND 20', '5', '0.012')
            thread = call('res2', '950000000', 'OUT 950.0 MOHM; ZCOMP WIRE2', 'CONF:RES 1 GOHM', 'C266', 'D266', 'DET:BAND 20', '5', '0.011')
        thread = message('Калибровка завершена')
        thread = reset()
    
    if a1[2] == '34401A':
        wb = load_workbook(shablon + '\\34401A.xlsx')
        ws = wb.active
        progress1.configure(maximum = 50)
        lb.insert(END, 'Время начала:' + ' ' + d1)
        # V
        thread = call('dc', '0.1', 'OUT 0.1 V', 'CONF:VOLT:DC 0.1', 'C28', 'D28', 'DET:BAND 20', '3', '0.009')
        thread = call('dc', '-0.1', 'OUT -0.1 V', 'CONF:VOLT:DC 0.1', 'C29', 'D29', 'DET:BAND 20', '3', '0.009')
        thread = call('dc', '1.0', 'OUT 1.0 V', 'CONF:VOLT:DC 1.0', 'C30', 'D30', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '-1.0', 'OUT -1.0 V', 'CONF:VOLT:DC 1.0', 'C31', 'D31', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '10.0', 'OUT 10.0 V', 'CONF:VOLT:DC 10', 'C32', 'D32', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '-10.0', 'OUT -10.0 V', 'CONF:VOLT:DC 10', 'C33', 'D33', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '100', 'OUT 100 V', 'CONF:VOLT:DC 100', 'C34', 'D34', 'DET:BAND 20', '4', '0.005')
        thread = call('dc', '-100', 'OUT -100 V', 'CONF:VOLT:DC 100', 'C35', 'D35', 'DET:BAND 20', '4', '0.005')
        thread = call('dc', '1000', 'OUT 1000 V', 'CONF:VOLT:DC 1000', 'C36', 'D36', 'DET:BAND 20', '4', '0.006')
        thread = reset()
        thread = call('dc', '-1000', 'OUT -1000 V', 'CONF:VOLT:DC 1000', 'C37', 'D37', 'DET:BAND 20', '8', '0.006')
        thread = reset()
        # ~V
        thread = call('ac', '10.0', 'OUT 10.0 V, 10 Hz', 'CONF:VOLT:AC 10.0', 'D42', 'E42', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '0.01', 'OUT 0.01 V, 1 kHz', 'CONF:VOLT:AC 0.1', 'D43', 'E43', 'DET:BAND 3', '8', '0.46')
        thread = call('ac', '0.1', 'OUT 0.1 V, 1 kHz', 'CONF:VOLT:AC 0.1', 'D44', 'E44', 'DET:BAND 3', '8', '0.1')
        thread = call('ac', '1.0', 'OUT 1.0 V, 1 kHz', 'CONF:VOLT:AC 1.0', 'D45', 'E45', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '10.0', 'OUT 10.0 V, 1 kHz', 'CONF:VOLT:AC 10.0', 'D46', 'E46', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '100.0', 'OUT 100.0 V, 1 kHz', 'CONF:VOLT:AC 100.0', 'D47', 'E47', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '750.0', 'OUT 750.0 V, 1 kHz', 'CONF:VOLT:AC 750.0', 'D48', 'E48', 'DET:BAND 3', '8', '0.083')
        thread = call('ac', '0.1', 'OUT 0.1 V, 50 kHz', 'CONF:VOLT:AC 0.1', 'D49', 'E49', 'DET:BAND 3', '8', '0.17')
        thread = call('ac', '1.0', 'OUT 1.0 V, 50 kHz', 'CONF:VOLT:AC 1.0', 'D50', 'E50', 'DET:BAND 3', '8', '0.17')
        thread = call('ac', '10.0', 'OUT 10.0 V, 50 kHz', 'CONF:VOLT:AC 10.0', 'D51', 'E51', 'DET:BAND 3', '8', '0.17')
        if b1[1] == '5522A':
            thread = call('ac', '100.0', 'OUT 100.0 V, 50 kHz', 'CONF:VOLT:AC 100.0', 'D52', 'E52', 'DET:BAND 20', '5', '0.17')
            #thread = call('ac', '750.0', 'OUT 750.0 V, 50 kHz', 'CONF:VOLT:AC 750.0', 'D53', 'E53', 'DET:BAND 20', '5', '0.17')
        thread = reset()
        # Hz
        thread = call('fr', '100.0', 'OUT 0.01 V, 100.0 Hz', 'CONF:FREQ 100.0 Hz', 'D74', 'E74', 'DET:BAND 20', '5', '0.1')
        thread = call('fr', '100000.0', 'OUT 1.0 V, 100.0 kHz', 'CONF:FREQ 100.0 kHz', 'D75', 'E75', 'DET:BAND 20', '5', '0.01')
        thread = message('Переключите провода\n для измерения тока')
        thread = reset()
        # A
        thread = call('dci', '0.01', 'OUT 0.01 A', 'CONF:CURR:DC 0.01', 'C58', 'D58', 'DET:BAND 20', '5', '0.07')
        thread = call('dci', '0.1', 'OUT 0.1 A', 'CONF:CURR:DC 0.1', 'C59', 'D59', 'DET:BAND 20', '5', '0.055')
        thread = call('dci', '1.0', 'OUT 1.0 A', 'CONF:CURR:DC 1.0', 'C60', 'D60', 'DET:BAND 20', '5', '0.11')
        thread = call('dci', '2.0', 'OUT 2.0 A', 'CONF:CURR:DC 3.0', 'C61', 'D61', 'DET:BAND 20', '5', '0.15')
        thread = reset()
        # ~A
        thread = call('aci', '1.0', 'OUT 1.0 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D66', 'E66', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '2.0', 'OUT 2.0 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D67', 'E67', 'DET:BAND 20', '5', '0.24')
        thread = message('Переключите провода по четырехпроводной схеме\n для измерения сопротивления')
        thread = reset()
        # Ohm
        thread = call('res4', '100', 'OUT 100 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'C80', 'D80', 'DET:BAND 20', '5', '0.014')
        thread = call('res4', '1000', 'OUT 1 kOHM; ZCOMP WIRE4', 'CONF:FRES 1 KOHM', 'C81', 'D81', 'DET:BAND 20', '5', '0.011')
        thread = call('res4', '10000', 'OUT 10 kOHM; ZCOMP WIRE4', 'CONF:FRES 10 KOHM', 'C82', 'D82', 'DET:BAND 20', '5', '0.011')
        thread = call('res4', '100000', 'OUT 100 kOHM; ZCOMP WIRE4', 'CONF:FRES 100 KOHM', 'C83', 'D83', 'DET:BAND 20', '5', '0.011')
        thread = message('Переключите провода по двухпроводной схеме\n для измерения сопротивления')
        thread = reset()
        thread = call('res2', '1000000', 'OUT 1 MOHM; ZCOMP WIRE2', 'CONF:RES 1 MOHM', 'C91', 'D91', 'DET:BAND 20', '5', '0.011')
        thread = call('res2', '10000000', 'OUT 10 MOHM; ZCOMP WIRE2', 'CONF:RES 10 MOHM', 'C92', 'D92', 'DET:BAND 20', '5', '0.041')
        thread = call('res2', '100000000', 'OUT 100 MOHM; ZCOMP WIRE2', 'CONF:RES 100 MOHM', 'C93', 'D93', 'DET:BAND 20', '5', '0.801')
        thread = message('Калибровка завершена')
        thread = reset()
        
    if a1[2] == '34460A':
        wb = load_workbook(shablon + '\\34460A.xlsx')
        ws = wb.active
        progress1.configure(maximum = 72)
        lb.insert(END, 'Время начала:' + ' ' + d1)
        # V
        thread = call('dc', '0.1', 'OUT 0.1 V', 'CONF:VOLT:DC 0.1', 'C28', 'D28', 'DET:BAND 20', '3', '0.016')
        thread = call('dc', '-0.1', 'OUT -0.1 V', 'CONF:VOLT:DC 0.1', 'C29', 'D29', 'DET:BAND 20', '3', '0.016')
        thread = call('dc', '1.0', 'OUT 1.0 V', 'CONF:VOLT:DC 1.0', 'C30', 'D30', 'DET:BAND 20', '3', '0.009')
        thread = call('dc', '-1.0', 'OUT -1.0 V', 'CONF:VOLT:DC 1.0', 'C31', 'D31', 'DET:BAND 20', '3', '0.009')
        thread = call('dc', '4.0', 'OUT 4.0 V', 'CONF:VOLT:DC 10', 'C32', 'D32', 'DET:BAND 20', '3', '0.009')
        thread = call('dc', '10.0', 'OUT 10.0 V', 'CONF:VOLT:DC 10', 'C33', 'D33', 'DET:BAND 20', '3', '0.008')
        thread = call('dc', '-10.0', 'OUT -10.0 V', 'CONF:VOLT:DC 10', 'C34', 'D34', 'DET:BAND 20', '3', '0.008')
        thread = call('dc', '100', 'OUT 100 V', 'CONF:VOLT:DC 100', 'C35', 'D35', 'DET:BAND 20', '4', '0.009')
        thread = call('dc', '-100', 'OUT -100 V', 'CONF:VOLT:DC 100', 'C36', 'D36', 'DET:BAND 20', '4', '0.009')
        thread = call('dc', '1000', 'OUT 1000 V', 'CONF:VOLT:DC 1000', 'C37', 'D37', 'DET:BAND 20', '8', '0.01')
        thread = reset()
        thread = call('dc', '-500', 'OUT -500 V', 'CONF:VOLT:DC 1000', 'C38', 'D38', 'DET:BAND 20', '8', '0.011')
        thread = reset()
        # ~V
        thread = call('ac', '10.0', 'OUT 10.0 V, 10 Hz', 'CONF:VOLT:AC 10.0', 'D43', 'E43', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '10.0', 'OUT 10.0 V, 100 Hz', 'CONF:VOLT:AC 10.0', 'D44', 'E44', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '0.1', 'OUT 0.1 V, 1 kHz', 'CONF:VOLT:AC 0.1', 'D45', 'E45', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '1.0', 'OUT 1.0 V, 1 kHz', 'CONF:VOLT:AC 1.0', 'D46', 'E46', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '0.03', 'OUT 0.03 V, 1 kHz', 'CONF:VOLT:AC 10.0', 'D47', 'E47', 'DET:BAND 20', '5', '10.0')
        thread = call('ac', '1.0', 'OUT 1.0 V, 1 kHz', 'CONF:VOLT:AC 10.0', 'D48', 'E48', 'DET:BAND 20', '5', '0.39')
        thread = call('ac', '100.0', 'OUT 100.0 V, 1 kHz', 'CONF:VOLT:AC 100.0', 'D49', 'E49', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '750.0', 'OUT 750.0 V, 1 kHz', 'CONF:VOLT:AC 750.0', 'D50', 'E50', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '10.0', 'OUT 10.0 V, 20 kHz', 'CONF:VOLT:AC 10.0', 'D51', 'E51', 'DET:BAND 20', '5', '0.12')
        thread = call('ac', '0.1', 'OUT 0.1 V, 50 kHz', 'CONF:VOLT:AC 0.1', 'D52', 'E52', 'DET:BAND 20', '5', '0.2')
        thread = call('ac', '1.0', 'OUT 1.0 V, 50 kHz', 'CONF:VOLT:AC 1.0', 'D53', 'E53', 'DET:BAND 20', '5', '0.2')
        thread = call('ac', '10.0', 'OUT 10.0 V, 50 kHz', 'CONF:VOLT:AC 10.0', 'D54', 'E54', 'DET:BAND 20', '5', '0.2')
        if b1[1] == '5522A':
            thread = call('ac', '100.0', 'OUT 100.0 V, 50 kHz', 'CONF:VOLT:AC 100.0', 'D55', 'E55', 'DET:BAND 20', '5', '0.2')
            thread = call('ac', '750.0', 'OUT 750.0 V, 50 kHz', 'CONF:VOLT:AC 750.0', 'D56', 'E56', 'DET:BAND 20', '5', '0.329')
        thread = call('ac', '10.0', 'OUT 10.0 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D57', 'E57', 'DET:BAND 20', '5', '0.71')
        thread = call('ac', '0.1', 'OUT 0.1 V, 300 kHz', 'CONF:VOLT:AC 0.1', 'D58', 'E58', 'DET:BAND 20', '5', '4.5')
        thread = call('ac', '1.0', 'OUT 1.0 V, 300 kHz', 'CONF:VOLT:AC 1.0', 'D59', 'E59', 'DET:BAND 20', '5', '4.5')
        thread = call('ac', '10.0', 'OUT 10.0 V, 300 kHz', 'CONF:VOLT:AC 10.0', 'D60', 'E60', 'DET:BAND 20', '5', '4.5')
        thread = call('ac', '70.0', 'OUT 70.0 V, 300 kHz', 'CONF:VOLT:AC 100.0', 'D61', 'E61', 'DET:BAND 20', '5', '4.714')
        thread = call('ac', '70.0', 'OUT 70.0 V, 300 kHz', 'CONF:VOLT:AC 750.0', 'D62', 'E62', 'DET:BAND 20', '5', '9.429')
        thread = reset()
        # Hz
        thread = call('fr', '10.0', 'OUT 0.1 V, 10.0 Hz', 'CONF:FREQ 10.0 Hz', 'D98', 'E98', 'DET:BAND 20', '5', '0.035')
        thread = call('fr', '300000.0', 'OUT 0.01 V, 300.0 kHz', 'CONF:FREQ 300.0 kHz', 'D99', 'E99', 'DET:BAND 20', '5', '0.17')
        thread = message('Переключите провода\n для измерения тока')
        thread = reset()
        # A
        thread = call('dci', '0.0001', 'OUT 0.0001 A', 'CONF:CURR:DC 0.0001', 'C67', 'D67', 'DET:BAND 20', '5', '0.075')
        thread = call('dci', '0.001', 'OUT 0.001 A', 'CONF:CURR:DC 0.001', 'C68', 'D68', 'DET:BAND 20', '5', '0.056')
        thread = call('dci', '0.01', 'OUT 0.01 A', 'CONF:CURR:DC 0.01', 'C69', 'D69', 'DET:BAND 20', '5', '0.07')
        thread = call('dci', '0.1', 'OUT 0.1 A', 'CONF:CURR:DC 0.1', 'C70', 'D70', 'DET:BAND 20', '5', '0.055')
        thread = call('dci', '1.0', 'OUT 1.0 A', 'CONF:CURR:DC 1.0', 'C71', 'D71', 'DET:BAND 20', '5', '0.11')
        thread = call('dci', '2.0', 'OUT 2.0 A', 'CONF:CURR:DC 3.0', 'C72', 'D72', 'DET:BAND 20', '5', '0.23')
        thread = reset()
        # ~A
        thread = call('aci', '0.1', 'OUT 0.1 A, 10 Hz', 'CONF:CURR:AC 0.1', 'D77', 'E77', 'DET:BAND 3', '8', '0.14')
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 1 kHz', 'CONF:CURR:AC 0.0001', 'D78', 'E78', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.001', 'OUT 0.001 A, 1 kHz', 'CONF:CURR:AC 0.001', 'D79', 'E79', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D80', 'E80', 'DET:BAND 20', '5', '4.1')
        thread = call('aci', '0.001', 'OUT 0.001 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D81', 'E81', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.01', 'OUT 0.01 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D82', 'E82', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.1', 'OUT 0.1 A, 1 kHz', 'CONF:CURR:AC 0.1', 'D83', 'E83', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '1.0', 'OUT 1.0 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D84', 'E84', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '2.0', 'OUT 2.0 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D85', 'E85', 'DET:BAND 20', '5', '0.29')
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 5 kHz', 'CONF:CURR:AC 0.0001', 'D86', 'E86', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.001', 'OUT 0.001 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D87', 'E87', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.01', 'OUT 0.01 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D88', 'E88', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.1', 'OUT 0.1 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D89', 'E89', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '1.0', 'OUT 1.0 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D90', 'E90', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '2.0', 'OUT 2.0 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D91', 'E91', 'DET:BAND 20', '5', '0.29')
        thread = message('Переключите провода по четырехпроводной схеме\n для измерения сопротивления')
        thread = reset()
        # Ohm
        thread = call('res4', '100', 'OUT 100 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'C104', 'D104', 'DET:BAND 20', '5', '0.021')
        thread = call('res4', '1000', 'OUT 1 kOHM; ZCOMP WIRE4', 'CONF:FRES 1 KOHM', 'C105', 'D105', 'DET:BAND 20', '5', '0.015')
        thread = call('res4', '10000', 'OUT 10 kOHM; ZCOMP WIRE4', 'CONF:FRES 10 KOHM', 'C106', 'D106', 'DET:BAND 20', '5', '0.015')
        thread = call('res4', '100000', 'OUT 100 kOHM; ZCOMP WIRE4', 'CONF:FRES 100 KOHM', 'C107', 'D107', 'DET:BAND 20', '5', '0.015')
        thread = message('Переключите провода по двухпроводной схеме\n для измерения сопротивления')
        thread = reset()
        thread = call('res2', '1000000', 'OUT 1 MOHM; ZCOMP WIRE2', 'CONF:RES 1 MOHM', 'C115', 'D115', 'DET:BAND 20', '5', '0.015')
        thread = call('res2', '10000000', 'OUT 10 MOHM; ZCOMP WIRE2', 'CONF:RES 10 MOHM', 'C116', 'D116', 'DET:BAND 20', '5', '0.041')
        thread = call('res2', '100000000', 'OUT 100 MOHM; ZCOMP WIRE2', 'CONF:RES 100 MOHM', 'C117', 'D117', 'DET:BAND 20', '5', '0.81')
        thread = message('Калибровка завершена')
        thread = reset()
        
    if a1[2] == '34461A':
        wb = load_workbook(shablon + '\\34461A.xlsx')
        ws = wb.active
        progress1.configure(maximum = 75)
        lb.insert(END, 'Время начала:' + ' ' + d1)
        # V
        thread = call('dc', '0.1', 'OUT 0.1 V', 'CONF:VOLT:DC 0.1', 'C28', 'D28', 'DET:BAND 20', '3', '0.009')
        thread = call('dc', '-0.1', 'OUT -0.1 V', 'CONF:VOLT:DC 0.1', 'C29', 'D29', 'DET:BAND 20', '3', '0.009')
        thread = call('dc', '1.0', 'OUT 1.0 V', 'CONF:VOLT:DC 1.0', 'C30', 'D30', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '-1.0', 'OUT -1.0 V', 'CONF:VOLT:DC 1.0', 'C31', 'D31', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '4.0', 'OUT 4.0 V', 'CONF:VOLT:DC 10', 'C32', 'D32', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '10.0', 'OUT 10.0 V', 'CONF:VOLT:DC 10', 'C33', 'D33', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '-10.0', 'OUT -10.0 V', 'CONF:VOLT:DC 10', 'C34', 'D34', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '100', 'OUT 100 V', 'CONF:VOLT:DC 100', 'C35', 'D35', 'DET:BAND 20', '4', '0.005')
        thread = call('dc', '-100', 'OUT -100 V', 'CONF:VOLT:DC 100', 'C36', 'D36', 'DET:BAND 20', '4', '0.005')
        thread = call('dc', '1000', 'OUT 1000 V', 'CONF:VOLT:DC 1000', 'C37', 'D37', 'DET:BAND 20', '4', '0.006')
        thread = reset()
        thread = call('dc', '-500', 'OUT -500 V', 'CONF:VOLT:DC 1000', 'C38', 'D38', 'DET:BAND 20', '8', '0.007')
        thread = reset()
        # ~V
        thread = call('ac', '10.0', 'OUT 10.0 V, 10 Hz', 'CONF:VOLT:AC 10.0', 'D43', 'E43', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '10.0', 'OUT 10.0 V, 100 Hz', 'CONF:VOLT:AC 10.0', 'D44', 'E44', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '0.1', 'OUT 0.1 V, 1 kHz', 'CONF:VOLT:AC 0.1', 'D45', 'E45', 'DET:BAND 20', '5', '0.09')
        thread = call('ac', '1.0', 'OUT 1.0 V, 1 kHz', 'CONF:VOLT:AC 1.0', 'D46', 'E46', 'DET:BAND 20', '5', '0.09')
        thread = call('ac', '0.03', 'OUT 0.03 V, 1 kHz', 'CONF:VOLT:AC 10.0', 'D47', 'E47', 'DET:BAND 20', '5', '10.0')
        thread = call('ac', '1.0', 'OUT 1.0 V, 1 kHz', 'CONF:VOLT:AC 10.0', 'D48', 'E48', 'DET:BAND 20', '5', '0.36')
        thread = call('ac', '100.0', 'OUT 100.0 V, 1 kHz', 'CONF:VOLT:AC 100.0', 'D49', 'E49', 'DET:BAND 20', '5', '0.09')
        thread = call('ac', '750.0', 'OUT 750.0 V, 1 kHz', 'CONF:VOLT:AC 750.0', 'D50', 'E50', 'DET:BAND 20', '5', '0.09')
        thread = call('ac', '10.0', 'OUT 10.0 V, 20 kHz', 'CONF:VOLT:AC 10.0', 'D51', 'E51', 'DET:BAND 20', '5', '0.09')
        thread = call('ac', '0.1', 'OUT 0.1 V, 50 kHz', 'CONF:VOLT:AC 0.1', 'D52', 'E52', 'DET:BAND 20', '5', '0.17')
        thread = call('ac', '1.0', 'OUT 1.0 V, 50 kHz', 'CONF:VOLT:AC 1.0', 'D53', 'E53', 'DET:BAND 20', '5', '0.17')
        thread = call('ac', '10.0', 'OUT 10.0 V, 50 kHz', 'CONF:VOLT:AC 10.0', 'D54', 'E54', 'DET:BAND 20', '5', '0.17')
        if b1[1] == '5522A':
            thread = call('ac', '100.0', 'OUT 100.0 V, 50 kHz', 'CONF:VOLT:AC 100.0', 'D55', 'E55', 'DET:BAND 20', '5', '0.17')
            thread = call('ac', '210.0', 'OUT 210.0 V, 50 kHz', 'CONF:VOLT:AC 750.0', 'D56', 'E56', 'DET:BAND 20', '5', '0.299')
        thread = call('ac', '10.0', 'OUT 10.0 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D57', 'E57', 'DET:BAND 20', '5', '0.68')
        thread = call('ac', '0.1', 'OUT 0.1 V, 300 kHz', 'CONF:VOLT:AC 0.1', 'D58', 'E58', 'DET:BAND 20', '5', '4.5')
        thread = call('ac', '1.0', 'OUT 1.0 V, 300 kHz', 'CONF:VOLT:AC 1.0', 'D59', 'E59', 'DET:BAND 20', '5', '4.5')
        thread = call('ac', '10.0', 'OUT 10.0 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D60', 'E60', 'DET:BAND 20', '5', '4.5')
        if b1[1] == '5522A':
            thread = call('ac', '70.0', 'OUT 70.0 V, 100 kHz', 'CONF:VOLT:AC 100.0', 'D61', 'E61', 'DET:BAND 20', '5', '4.714')
            thread = call('ac', '70.0', 'OUT 70.0 V, 100 kHz', 'CONF:VOLT:AC 750.0', 'D62', 'E62', 'DET:BAND 20', '5', '9.429')
        thread = reset()
        # Hz
        thread = call('fr', '10.0', 'OUT 0.1 V, 10.0 Hz', 'CONF:FREQ 10.0 Hz', 'D101', 'E101', 'DET:BAND 20', '5', '0.035')
        thread = call('fr', '300000.0', 'OUT 0.01 V, 300.0 kHz', 'CONF:FREQ 300.0 kHz', 'D102', 'E102', 'DET:BAND 20', '5', '0.17')
        thread = message('Переключите провода\n для измерения тока')
        thread = reset()
        # A
        thread = call('dci', '0.0001', 'OUT 0.0001 A', 'CONF:CURR:DC 0.0001', 'C67', 'D67', 'DET:BAND 20', '5', '0.075')
        thread = call('dci', '0.001', 'OUT 0.001 A', 'CONF:CURR:DC 0.001', 'C68', 'D68', 'DET:BAND 20', '5', '0.056')
        thread = call('dci', '0.01', 'OUT 0.01 A', 'CONF:CURR:DC 0.01', 'C69', 'D69', 'DET:BAND 20', '5', '0.07')
        thread = call('dci', '0.1', 'OUT 0.1 A', 'CONF:CURR:DC 0.1', 'C70', 'D70', 'DET:BAND 20', '5', '0.055')
        thread = call('dci', '1.0', 'OUT 1.0 A', 'CONF:CURR:DC 1.0', 'C71', 'D71', 'DET:BAND 20', '5', '0.11')
        thread = call('dci', '2.0', 'OUT 2.0 A', 'CONF:CURR:DC 3.0', 'C72', 'D72', 'DET:BAND 20', '5', '0.23')
        thread = reset()
        # ~A
        thread = call('aci', '0.1', 'OUT 0.1 A, 10 Hz', 'CONF:CURR:AC 0.1', 'D79', 'E79', 'DET:BAND 3', '8', '0.14')
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 1 kHz', 'CONF:CURR:AC 0.0001', 'D80', 'E80', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.001', 'OUT 0.001 A, 1 kHz', 'CONF:CURR:AC 0.001', 'D81', 'E81', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D82', 'E82', 'DET:BAND 20', '5', '4.1')
        thread = call('aci', '0.001', 'OUT 0.001 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D83', 'E83', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.01', 'OUT 0.01 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D84', 'E84', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.1', 'OUT 0.1 A, 1 kHz', 'CONF:CURR:AC 0.1', 'D85', 'E85', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '1.0', 'OUT 1.0 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D86', 'E86', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '2.0', 'OUT 2.0 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D87', 'E87', 'DET:BAND 20', '5', '0.29')
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 5 kHz', 'CONF:CURR:AC 0.0001', 'D88', 'E88', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.001', 'OUT 0.001 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D89', 'E89', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.01', 'OUT 0.01 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D90', 'E90', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '0.1', 'OUT 0.1 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D91', 'E91', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '1.0', 'OUT 1.0 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D92', 'E92', 'DET:BAND 20', '5', '0.14')
        thread = call('aci', '2.0', 'OUT 2.0 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D93', 'E93', 'DET:BAND 20', '5', '0.29')
        # 10А
        thread = message('Переключите красный провод на контакт 10А мультиметра')
        thread = reset()
        if b1[1] == '5500E':
            thread = call('dci', '5.0', 'OUT 5.0 A', 'CONF:CURR:DC 10.0', 'C73', 'D73', 'DET:BAND 20', '5', '0.24')
            thread = call('dci', '10.0', 'OUT 10.0 A', 'CONF:CURR:DC 10.0', 'C74', 'D74', 'DET:BAND 20', '5', '0.22')
        if b1[1] == '5522A':
            thread = message('Переключите красный провод на калибраторе в разъем больше 2,5 А')
            thread = call('dci', '5.0', 'OUT 5.0 A', 'CONF:CURR:DC 10.0', 'C73', 'D73', 'DET:BAND 20', '5', '0.24')
            thread = call('dci', '10.0', 'OUT 10.0 A', 'CONF:CURR:DC 10.0', 'C74', 'D74', 'DET:BAND 20', '5', '0.22')
            thread = call('aci', '10.0', 'OUT 10.0 A, 5 kHz', 'CONF:CURR:AC 10.0', 'D94', 'E94', 'DET:BAND 20', '5', '0.19')
        thread = message('Переключите провода по четырехпроводной схеме\n для измерения сопротивления')
        thread = reset()
        # Ohm
        thread = call('res4', '100', 'OUT 100 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'C107', 'D107', 'DET:BAND 20', '5', '0.014')
        thread = call('res4', '1000', 'OUT 1 kOHM; ZCOMP WIRE4', 'CONF:FRES 1 KOHM', 'C108', 'D108', 'DET:BAND 20', '5', '0.011')
        thread = call('res4', '10000', 'OUT 10 kOHM; ZCOMP WIRE4', 'CONF:FRES 10 KOHM', 'C109', 'D109', 'DET:BAND 20', '5', '0.011')
        thread = call('res4', '100000', 'OUT 100 kOHM; ZCOMP WIRE4', 'CONF:FRES 100 KOHM', 'C110', 'D110', 'DET:BAND 20', '5', '0.011')
        thread = message('Переключите провода по двухпроводной схеме\n для измерения сопротивления')
        thread = reset()
        thread = call('res2', '1000000', 'OUT 1 MOHM; ZCOMP WIRE2', 'CONF:RES 1 MOHM', 'C118', 'D118', 'DET:BAND 20', '5', '0.011')
        thread = call('res2', '10000000', 'OUT 10 MOHM; ZCOMP WIRE2', 'CONF:RES 10 MOHM', 'C119', 'D119', 'DET:BAND 20', '5', '0.041')
        thread = call('res2', '100000000', 'OUT 100 MOHM; ZCOMP WIRE2', 'CONF:RES 100 MOHM', 'C120', 'D120', 'DET:BAND 20', '5', '0.81')
        thread = message('Калибровка завершена')
        thread = reset()
        
    if a1[2] in ('34465A', '34470A'):
        wb = load_workbook(shablon + '\\34465A,34470A.xlsx')
        ws = wb.active
        progress1.configure(maximum = 280)
        lb.insert(END, 'Время начала:' + ' ' + d1)
        # 0.1V
        thread = call('dc', '0.01', 'OUT 0.01 V', 'CONF:VOLT:DC 0.1', 'C28', 'D28', 'DET:BAND 20', '3', '0.04')
        thread = call('dc', '0.03', 'OUT 0.03 V', 'CONF:VOLT:DC 0.1', 'C29', 'D29', 'DET:BAND 20', '3', '0.017')
        thread = call('dc', '0.05', 'OUT 0.05 V', 'CONF:VOLT:DC 0.1', 'C30', 'D30', 'DET:BAND 20', '3', '0.012')
        thread = call('dc', '0.07', 'OUT 0.07 V', 'CONF:VOLT:DC 0.1', 'C31', 'D31', 'DET:BAND 20', '3', '0.01')
        thread = call('dc', '0.1', 'OUT 0.1 V', 'CONF:VOLT:DC 0.1', 'C32', 'D32', 'DET:BAND 20', '3', '0.009')
        # 1V
        thread = call('dc', '0.1', 'OUT 0.1 V', 'CONF:VOLT:DC 1.0', 'C33', 'D33', 'DET:BAND 20', '3', '0.008')
        thread = call('dc', '0.3', 'OUT 0.3 V', 'CONF:VOLT:DC 1.0', 'C34', 'D34', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '0.5', 'OUT 0.5 V', 'CONF:VOLT:DC 1.0', 'C35', 'D35', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '0.7', 'OUT 0.7 V', 'CONF:VOLT:DC 1.0', 'C36', 'D36', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '1.0', 'OUT 1.0 V', 'CONF:VOLT:DC 1.0', 'C37', 'D37', 'DET:BAND 20', '3', '0.004')
        # 10V
        thread = call('dc', '1.0', 'OUT 1.0 V', 'CONF:VOLT:DC 10', 'C38', 'D38', 'DET:BAND 20', '3', '0.007')
        thread = call('dc', '3.0', 'OUT 3.0 V', 'CONF:VOLT:DC 10', 'C39', 'D39', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '5.0', 'OUT 5.0 V', 'CONF:VOLT:DC 10', 'C40', 'D40', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '7.0', 'OUT 7.0 V', 'CONF:VOLT:DC 10', 'C41', 'D41', 'DET:BAND 20', '3', '0.004')
        thread = call('dc', '10.0', 'OUT 10.0 V', 'CONF:VOLT:DC 10', 'C42', 'D42', 'DET:BAND 20', '3', '0.003')
        # 100V
        thread = call('dc', '10', 'OUT 10 V', 'CONF:VOLT:DC 100', 'C43', 'D43', 'DET:BAND 20', '3', '0.01')
        thread = call('dc', '30', 'OUT 30 V', 'CONF:VOLT:DC 100', 'C44', 'D44', 'DET:BAND 20', '3', '0.006')
        thread = call('dc', '50', 'OUT 50 V', 'CONF:VOLT:DC 100', 'C45', 'D45', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '70', 'OUT 70 V', 'CONF:VOLT:DC 100', 'C46', 'D46', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '100', 'OUT 100 V', 'CONF:VOLT:DC 100', 'C47', 'D47', 'DET:BAND 20', '3', '0.005')
        # 1000V
        thread = call('dc', '100', 'OUT 100 V', 'CONF:VOLT:DC 1000', 'C48', 'D48', 'DET:BAND 20', '3', '0.01')
        thread = call('dc', '300', 'OUT 300 V', 'CONF:VOLT:DC 1000', 'C49', 'D49', 'DET:BAND 20', '3', '0.006')
        thread = call('dc', '500', 'OUT 500 V', 'CONF:VOLT:DC 1000', 'C50', 'D50', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '700', 'OUT 700 V', 'CONF:VOLT:DC 1000', 'C51', 'D51', 'DET:BAND 20', '3', '0.005')
        thread = call('dc', '1000', 'OUT 1000 V', 'CONF:VOLT:DC 1000', 'C52', 'D52', 'DET:BAND 20', '3', '0.006')
        thread = reset()
        # ~0.1V,20Hz
        thread = call('ac', '0.01', 'OUT 0.01 V, 20 Hz', 'CONF:VOLT:AC 0.1', 'D57', 'E57', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '0.03', 'OUT 0.03 V, 20 Hz', 'CONF:VOLT:AC 0.1', 'D58', 'E58', 'DET:BAND 3', '8', '0.117')
        thread = call('ac', '0.05', 'OUT 0.05 V, 20 Hz', 'CONF:VOLT:AC 0.1', 'D59', 'E59', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '0.07', 'OUT 0.07 V, 20 Hz', 'CONF:VOLT:AC 0.1', 'D60', 'E60', 'DET:BAND 3', '8', '0.079')
        thread = call('ac', '0.1', 'OUT 0.1 V, 20 Hz', 'CONF:VOLT:AC 0.1', 'D61', 'E61', 'DET:BAND 3', '8', '0.07')
        # ~1V,20Hz
        thread = call('ac', '0.1', 'OUT 0.1 V, 20 Hz', 'CONF:VOLT:AC 1.0', 'D62', 'E62', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '0.3', 'OUT 0.3 V, 20 Hz', 'CONF:VOLT:AC 1.0', 'D63', 'E63', 'DET:BAND 3', '8', '0.117')
        thread = call('ac', '0.5', 'OUT 0.5 V, 20 Hz', 'CONF:VOLT:AC 1.0', 'D64', 'E64', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '0.7', 'OUT 0.7 V, 20 Hz', 'CONF:VOLT:AC 1.0', 'D65', 'E65', 'DET:BAND 3', '8', '0.079')
        thread = call('ac', '1.0', 'OUT 1.0 V, 20 Hz', 'CONF:VOLT:AC 1.0', 'D66', 'E66', 'DET:BAND 3', '8', '0.07')
        # ~10V,20Hz
        thread = call('ac', '1', 'OUT 1 V, 20 Hz', 'CONF:VOLT:AC 10.0', 'D67', 'E67', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '3', 'OUT 3 V, 20 Hz', 'CONF:VOLT:AC 10.0', 'D68', 'E68', 'DET:BAND 3', '8', '0.117')
        thread = call('ac', '5', 'OUT 5 V, 20 Hz', 'CONF:VOLT:AC 10.0', 'D69', 'E69', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '7', 'OUT 7 V, 20 Hz', 'CONF:VOLT:AC 10.0', 'D70', 'E70', 'DET:BAND 3', '8', '0.079')
        thread = call('ac', '10', 'OUT 10 V, 20 Hz', 'CONF:VOLT:AC 10.0', 'D71', 'E71', 'DET:BAND 3', '8', '0.07')
        # ~100V,20Hz
        thread = call('ac', '10', 'OUT 10 V, 20 Hz', 'CONF:VOLT:AC 100', 'D72', 'E72', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '30', 'OUT 30 V, 20 Hz', 'CONF:VOLT:AC 100', 'D73', 'E73', 'DET:BAND 3', '8', '0.117')
        '''if b1[1] in ('5500A', '5522A'):
            thread = call('ac', '50', 'OUT 50 V, 20 Hz', 'CONF:VOLT:AC 100', 'D74', 'E74', 'DET:BAND 3', '8', '0.09')
            thread = call('ac', '70', 'OUT 70 V, 20 Hz', 'CONF:VOLT:AC 100', 'D75', 'E75', 'DET:BAND 3', '8', '0.079')
            thread = call('ac', '100', 'OUT 100 V, 20 Hz', 'CONF:VOLT:AC 100', 'D76', 'E76', 'DET:BAND 3', '8', '0.07')'''
        # ~0.1V,15kHz
        thread = call('ac', '0.01', 'OUT 0.01 V, 15 kHz', 'CONF:VOLT:AC 0.1', 'D77', 'E77', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '0.03', 'OUT 0.03 V, 15 kHz', 'CONF:VOLT:AC 0.1', 'D78', 'E78', 'DET:BAND 3', '8', '0.117')
        thread = call('ac', '0.05', 'OUT 0.05 V, 15 kHz', 'CONF:VOLT:AC 0.1', 'D79', 'E79', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '0.07', 'OUT 0.07 V, 15 kHz', 'CONF:VOLT:AC 0.1', 'D80', 'E80', 'DET:BAND 3', '8', '0.079')
        thread = call('ac', '0.1', 'OUT 0.1 V, 15 kHz', 'CONF:VOLT:AC 0.1', 'D81', 'E81', 'DET:BAND 3', '8', '0.07')
        # ~1V,15kHz
        thread = call('ac', '0.1', 'OUT 0.1 V, 15 kHz', 'CONF:VOLT:AC 1.0', 'D82', 'E82', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '0.3', 'OUT 0.3 V, 15 kHz', 'CONF:VOLT:AC 1.0', 'D83', 'E83', 'DET:BAND 3', '8', '0.117')
        thread = call('ac', '0.5', 'OUT 0.5 V, 15 kHz', 'CONF:VOLT:AC 1.0', 'D84', 'E84', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '0.7', 'OUT 0.7 V, 15 kHz', 'CONF:VOLT:AC 1.0', 'D85', 'E85', 'DET:BAND 3', '8', '0.079')
        thread = call('ac', '1.0', 'OUT 1.0 V, 15 kHz', 'CONF:VOLT:AC 1.0', 'D86', 'E86', 'DET:BAND 3', '8', '0.07')
        # ~10V,15kHz
        thread = call('ac', '1', 'OUT 1 V, 15 kHz', 'CONF:VOLT:AC 10.0', 'D87', 'E87', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '3', 'OUT 3 V, 15 kHz', 'CONF:VOLT:AC 10.0', 'D88', 'E88', 'DET:BAND 3', '8', '0.117')
        thread = call('ac', '5', 'OUT 5 V, 15 kHz', 'CONF:VOLT:AC 10.0', 'D89', 'E89', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '7', 'OUT 7 V, 15 kHz', 'CONF:VOLT:AC 10.0', 'D90', 'E90', 'DET:BAND 3', '8', '0.079')
        thread = call('ac', '10', 'OUT 10 V, 15 kHz', 'CONF:VOLT:AC 10.0', 'D91', 'E91', 'DET:BAND 3', '8', '0.07')
        # ~100V,15kHz
        thread = call('ac', '10', 'OUT 10 V, 15 kHz', 'CONF:VOLT:AC 100', 'D92', 'E92', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '30', 'OUT 30 V, 15 kHz', 'CONF:VOLT:AC 100', 'D93', 'E93', 'DET:BAND 3', '8', '0.117')
        thread = call('ac', '50', 'OUT 50 V, 15 kHz', 'CONF:VOLT:AC 100', 'D94', 'E94', 'DET:BAND 3', '8', '0.09')
        thread = call('ac', '70', 'OUT 70 V, 15 kHz', 'CONF:VOLT:AC 100', 'D95', 'E95', 'DET:BAND 3', '8', '0.079')
        thread = call('ac', '100', 'OUT 100 V, 15 kHz', 'CONF:VOLT:AC 100', 'D96', 'E96', 'DET:BAND 3', '8', '0.07')
        # ~750V,10kHz
        if b1[1] in ('5500A', '5522A'):
            thread = call('ac', '75', 'OUT 75 V, 10 kHz', 'CONF:VOLT:AC 750', 'D97', 'E97', 'DET:BAND 3', '8', '0.25')
            thread = call('ac', '225', 'OUT 225 V, 10 kHz', 'CONF:VOLT:AC 750', 'D98', 'E98', 'DET:BAND 3', '8', '0.117')
            thread = call('ac', '375', 'OUT 375 V, 10 kHz', 'CONF:VOLT:AC 750', 'D99', 'E99', 'DET:BAND 3', '8', '0.09')
            thread = call('ac', '525', 'OUT 525 V, 10 kHz', 'CONF:VOLT:AC 750', 'D100', 'E100', 'DET:BAND 3', '8', '0.121')
            thread = call('ac', '750', 'OUT 750 V, 10 kHz', 'CONF:VOLT:AC 750', 'D101', 'E101', 'DET:BAND 3', '8', '0.13')	
        # ~0.1V,30kHz
        thread = call('ac', '0.01', 'OUT 0.01 V, 30 kHz', 'CONF:VOLT:AC 0.1', 'D102', 'E102', 'DET:BAND 3', '8', '0.37')
        thread = call('ac', '0.03', 'OUT 0.03 V, 30 kHz', 'CONF:VOLT:AC 0.1', 'D103', 'E103', 'DET:BAND 3', '8', '0.17')
        thread = call('ac', '0.05', 'OUT 0.05 V, 30 kHz', 'CONF:VOLT:AC 0.1', 'D104', 'E104', 'DET:BAND 3', '8', '0.13')
        thread = call('ac', '0.07', 'OUT 0.07 V, 30 kHz', 'CONF:VOLT:AC 0.1', 'D105', 'E105', 'DET:BAND 3', '8', '0.113')
        thread = call('ac', '0.1', 'OUT 0.1 V, 30 kHz', 'CONF:VOLT:AC 0.1', 'D106', 'E106', 'DET:BAND 3', '8', '0.1')
        # ~1V,30kHz
        thread = call('ac', '0.1', 'OUT 0.1 V, 30 kHz', 'CONF:VOLT:AC 1.0', 'D107', 'E107', 'DET:BAND 3', '8', '0.37')
        thread = call('ac', '0.3', 'OUT 0.3 V, 30 kHz', 'CONF:VOLT:AC 1.0', 'D108', 'E108', 'DET:BAND 3', '8', '0.17')
        thread = call('ac', '0.5', 'OUT 0.5 V, 30 kHz', 'CONF:VOLT:AC 1.0', 'D109', 'E109', 'DET:BAND 3', '8', '0.13')
        thread = call('ac', '0.7', 'OUT 0.7 V, 30 kHz', 'CONF:VOLT:AC 1.0', 'D110', 'E110', 'DET:BAND 3', '8', '0.113')
        thread = call('ac', '1.0', 'OUT 1.0 V, 30 kHz', 'CONF:VOLT:AC 1.0', 'D111', 'E111', 'DET:BAND 3', '8', '0.1')
        # ~10V,30kHz
        thread = call('ac', '1', 'OUT 1 V, 30 kHz', 'CONF:VOLT:AC 10.0', 'D112', 'E112', 'DET:BAND 3', '8', '0.37')
        thread = call('ac', '3', 'OUT 3 V, 30 kHz', 'CONF:VOLT:AC 10.0', 'D113', 'E113', 'DET:BAND 3', '8', '0.17')
        thread = call('ac', '5', 'OUT 5 V, 30 kHz', 'CONF:VOLT:AC 10.0', 'D114', 'E114', 'DET:BAND 3', '8', '0.13')
        thread = call('ac', '7', 'OUT 7 V, 30 kHz', 'CONF:VOLT:AC 10.0', 'D115', 'E115', 'DET:BAND 3', '8', '0.113')
        thread = call('ac', '10', 'OUT 10 V, 30 kHz', 'CONF:VOLT:AC 10.0', 'D116', 'E116', 'DET:BAND 3', '8', '0.1')
        # ~100V,30kHz
        if b1[1] == '5522A':
            thread = call('ac', '10', 'OUT 10 V, 30 kHz', 'CONF:VOLT:AC 100', 'D117', 'E117', 'DET:BAND 3', '8', '0.37')
            thread = call('ac', '30', 'OUT 30 V, 30 kHz', 'CONF:VOLT:AC 100', 'D118', 'E118', 'DET:BAND 3', '8', '0.17')
            thread = call('ac', '50', 'OUT 50 V, 30 kHz', 'CONF:VOLT:AC 100', 'D119', 'E119', 'DET:BAND 3', '8', '0.13')
            thread = call('ac', '70', 'OUT 70 V, 30 kHz', 'CONF:VOLT:AC 100', 'D120', 'E120', 'DET:BAND 3', '8', '0.113')
            thread = call('ac', '100', 'OUT 100 V, 30 kHz', 'CONF:VOLT:AC 100', 'D121', 'E121', 'DET:BAND 3', '8', '0.1')
        # ~750V,30kHz
        if b1[1] == '5522A':
            thread = call('ac', '75', 'OUT 75 V, 30 kHz', 'CONF:VOLT:AC 750', 'D122', 'E122', 'DET:BAND 3', '8', '0.37')
            thread = call('ac', '225', 'OUT 225 V, 30 kHz', 'CONF:VOLT:AC 750', 'D123', 'E123', 'DET:BAND 3', '8', '0.17')
            '''thread = call('ac', '375', 'OUT 375 V, 30 kHz', 'CONF:VOLT:AC 750', 'D124', 'E124', 'DET:BAND 3', '8', '0.13')
            thread = call('ac', '525', 'OUT 525 V, 30 kHz', 'CONF:VOLT:AC 750', 'D125', 'E125', 'DET:BAND 3', '8', '0.156')
            thread = call('ac', '750', 'OUT 750 V, 30 kHz', 'CONF:VOLT:AC 750', 'D126', 'E126', 'DET:BAND 3', '8', '0.16')'''
        # ~0.1V,70kHz
        thread = call('ac', '0.01', 'OUT 0.01 V, 70 kHz', 'CONF:VOLT:AC 0.1', 'D127', 'E127', 'DET:BAND 3', '8', '0.65')
        thread = call('ac', '0.03', 'OUT 0.03 V, 70 kHz', 'CONF:VOLT:AC 0.1', 'D128', 'E128', 'DET:BAND 3', '8', '0.317')
        thread = call('ac', '0.05', 'OUT 0.05 V, 70 kHz', 'CONF:VOLT:AC 0.1', 'D129', 'E129', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '0.07', 'OUT 0.07 V, 70 kHz', 'CONF:VOLT:AC 0.1', 'D130', 'E130', 'DET:BAND 3', '8', '0.221')
        thread = call('ac', '0.1', 'OUT 0.1 V, 70 kHz', 'CONF:VOLT:AC 0.1', 'D131', 'E131', 'DET:BAND 3', '8', '0.2')
        # ~1V,70kHz
        thread = call('ac', '0.1', 'OUT 0.1 V, 70 kHz', 'CONF:VOLT:AC 1.0', 'D132', 'E132', 'DET:BAND 3', '8', '0.65')
        thread = call('ac', '0.3', 'OUT 0.3 V, 70 kHz', 'CONF:VOLT:AC 1.0', 'D133', 'E133', 'DET:BAND 3', '8', '0.317')
        thread = call('ac', '0.5', 'OUT 0.5 V, 70 kHz', 'CONF:VOLT:AC 1.0', 'D134', 'E134', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '0.7', 'OUT 0.7 V, 70 kHz', 'CONF:VOLT:AC 1.0', 'D135', 'E135', 'DET:BAND 3', '8', '0.221')
        thread = call('ac', '1.0', 'OUT 1.0 V, 70 kHz', 'CONF:VOLT:AC 1.0', 'D136', 'E136', 'DET:BAND 3', '8', '0.2')
        # ~10V,70kHz
        thread = call('ac', '1', 'OUT 1 V, 70 kHz', 'CONF:VOLT:AC 10.0', 'D137', 'E137', 'DET:BAND 3', '8', '0.65')
        thread = call('ac', '3', 'OUT 3 V, 70 kHz', 'CONF:VOLT:AC 10.0', 'D138', 'E138', 'DET:BAND 3', '8', '0.317')
        thread = call('ac', '5', 'OUT 5 V, 70 kHz', 'CONF:VOLT:AC 10.0', 'D139', 'E139', 'DET:BAND 3', '8', '0.25')
        thread = call('ac', '7', 'OUT 7 V, 70 kHz', 'CONF:VOLT:AC 10.0', 'D140', 'E140', 'DET:BAND 3', '8', '0.221')
        thread = call('ac', '10', 'OUT 10 V, 70 kHz', 'CONF:VOLT:AC 10.0', 'D141', 'E141', 'DET:BAND 3', '8', '0.2')
        # ~100V,70kHz
        if b1[1] == '5522A':
            thread = call('ac', '10', 'OUT 10 V, 70 kHz', 'CONF:VOLT:AC 100', 'D142', 'E142', 'DET:BAND 3', '8', '0.65')
            thread = call('ac', '30', 'OUT 30 V, 70 kHz', 'CONF:VOLT:AC 100', 'D143', 'E143', 'DET:BAND 3', '8', '0.317')
            thread = call('ac', '50', 'OUT 50 V, 70 kHz', 'CONF:VOLT:AC 100', 'D144', 'E144', 'DET:BAND 3', '8', '0.25')
            thread = call('ac', '70', 'OUT 70 V, 70 kHz', 'CONF:VOLT:AC 100', 'D145', 'E145', 'DET:BAND 3', '8', '0.221')
            thread = call('ac', '100', 'OUT 100 V, 70 kHz', 'CONF:VOLT:AC 100', 'D146', 'E146', 'DET:BAND 3', '8', '0.2')
        # ~750V,70kHz
        if b1[1] == '5522A':
            thread = call('ac', '75', 'OUT 75 V, 70 kHz', 'CONF:VOLT:AC 750', 'D147', 'E147', 'DET:BAND 3', '8', '0.65')
            thread = call('ac', '225', 'OUT 225 V, 70 kHz', 'CONF:VOLT:AC 750', 'D148', 'E148', 'DET:BAND 3', '8', '0.317')
            '''thread = call('ac', '375', 'OUT 375 V, 70 kHz', 'CONF:VOLT:AC 750', 'D149', 'E149', 'DET:BAND 3', '8', '0.25')
            thread = call('ac', '525', 'OUT 525 V, 70 kHz', 'CONF:VOLT:AC 750', 'D150', 'E150', 'DET:BAND 3', '8', '0.264')
            thread = call('ac', '750', 'OUT 750 V, 70 kHz', 'CONF:VOLT:AC 750', 'D151', 'E151', 'DET:BAND 3', '8', '0.26')'''
        # ~0.1V,200kHz
        thread = call('ac', '0.01', 'OUT 0.01 V, 200 kHz', 'CONF:VOLT:AC 0.1', 'D152', 'E152', 'DET:BAND 3', '8', '2')
        thread = call('ac', '0.03', 'OUT 0.03 V, 200 kHz', 'CONF:VOLT:AC 0.1', 'D153', 'E153', 'DET:BAND 3', '8', '1.333')
        thread = call('ac', '0.05', 'OUT 0.05 V, 200 kHz', 'CONF:VOLT:AC 0.1', 'D154', 'E154', 'DET:BAND 3', '8', '1.2')
        thread = call('ac', '0.07', 'OUT 0.07 V, 200 kHz', 'CONF:VOLT:AC 0.1', 'D155', 'E155', 'DET:BAND 3', '8', '1.143')
        thread = call('ac', '0.1', 'OUT 0.1 V, 200 kHz', 'CONF:VOLT:AC 0.1', 'D156', 'E156', 'DET:BAND 3', '8', '1.1')
        # ~1V,200kHz
        thread = call('ac', '0.1', 'OUT 0.1 V, 200 kHz', 'CONF:VOLT:AC 1.0', 'D157', 'E157', 'DET:BAND 3', '8', '2')
        thread = call('ac', '0.3', 'OUT 0.3 V, 200 kHz', 'CONF:VOLT:AC 1.0', 'D158', 'E158', 'DET:BAND 3', '8', '1.333')
        thread = call('ac', '0.5', 'OUT 0.5 V, 200 kHz', 'CONF:VOLT:AC 1.0', 'D159', 'E159', 'DET:BAND 3', '8', '1.2')
        thread = call('ac', '0.7', 'OUT 0.7 V, 200 kHz', 'CONF:VOLT:AC 1.0', 'D160', 'E160', 'DET:BAND 3', '8', '1.143')
        thread = call('ac', '1.0', 'OUT 1.0 V, 200 kHz', 'CONF:VOLT:AC 1.0', 'D161', 'E161', 'DET:BAND 3', '8', '1.1')
        # ~10V,100kHz
        thread = call('ac', '1', 'OUT 1 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D162', 'E162', 'DET:BAND 3', '8', '2')
        thread = call('ac', '3', 'OUT 3 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D163', 'E163', 'DET:BAND 3', '8', '1.333')
        thread = call('ac', '5', 'OUT 5 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D164', 'E164', 'DET:BAND 3', '8', '1.2')
        thread = call('ac', '7', 'OUT 7 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D165', 'E165', 'DET:BAND 3', '8', '1.143')
        thread = call('ac', '10', 'OUT 10 V, 100 kHz', 'CONF:VOLT:AC 10.0', 'D166', 'E166', 'DET:BAND 3', '8', '1.1')
        # ~100V,200kHz
        thread = call('ac', '10', 'OUT 10 V, 100 kHz', 'CONF:VOLT:AC 100', 'D167', 'E167', 'DET:BAND 3', '8', '2')
        thread = call('ac', '30', 'OUT 30 V, 100 kHz', 'CONF:VOLT:AC 100', 'D168', 'E168', 'DET:BAND 3', '8', '1.333')
        if b1[1] == '5522A':
            thread = call('ac', '50', 'OUT 50 V, 100 kHz', 'CONF:VOLT:AC 100', 'D169', 'E169', 'DET:BAND 3', '8', '1.2')
            thread = call('ac', '70', 'OUT 70 V, 100 kHz', 'CONF:VOLT:AC 100', 'D170', 'E170', 'DET:BAND 3', '8', '1.143')
            thread = call('ac', '100', 'OUT 100 V, 100 kHz', 'CONF:VOLT:AC 100', 'D171', 'E171', 'DET:BAND 3', '8', '1.1')
        # 5Hz
        thread = call('fr', '5.0', 'OUT 0.1 V, 5 Hz', 'CONF:FREQ 5.0 Hz', 'C299', 'D299', 'DET:BAND 20', '5', '0.07')
        thread = call('fr', '5.0', 'OUT 1.0 V, 5 Hz', 'CONF:FREQ 5.0 Hz', 'C300', 'D300', 'DET:BAND 20', '5', '0.07')
        # 50Hz
        thread = call('fr', '50.0', 'OUT 0.1 V, 50 Hz', 'CONF:FREQ 50.0 Hz', 'C301', 'D301', 'DET:BAND 20', '5', '0.03')
        thread = call('fr', '50.0', 'OUT 1.0 V, 50 Hz', 'CONF:FREQ 50.0 Hz', 'C302', 'D302', 'DET:BAND 20', '5', '0.03')
        # 500Hz
        thread = call('fr', '500.0', 'OUT 0.1 V, 500 Hz', 'CONF:FREQ 500.0 Hz', 'C303', 'D303', 'DET:BAND 20', '5', '0.007')
        thread = call('fr', '500.0', 'OUT 1.0 V, 500 Hz', 'CONF:FREQ 500.0 Hz', 'C304', 'D304', 'DET:BAND 20', '5', '0.007')
        # 100kHz
        thread = call('fr', '100000.0', 'OUT 0.1 V, 100 kHz', 'CONF:FREQ 100.0 kHz', 'C305', 'D305', 'DET:BAND 20', '5', '0.007')
        thread = call('fr', '100000.0', 'OUT 1.0 V, 100 kHz', 'CONF:FREQ 100.0 kHz', 'C306', 'D306', 'DET:BAND 20', '5', '0.007')
        # C
        thread = message('Измерение ёмкости.\nВытащите красный провод из каллибратора\nдля компенсации проводов')
        thread = cap()
        thread = message('Верните провод на место')
        thread = call('cap', '1E-9', 'OUT 1 NF', 'CONF:CAP 1 NF', 'C331', 'D331', 'DET:BAND 20', '5', '1.0')
        thread = call('cap', '10E-9', 'OUT 10 NF', 'CONF:CAP 10 NF', 'C332', 'D332', 'DET:BAND 20', '5', '0.5')
        thread = call('cap', '100E-9', 'OUT 100 NF', 'CONF:CAP 100 NF', 'C333', 'D333', 'DET:BAND 20', '5', '0.5')
        thread = call('cap', '1E-6', 'OUT 1 UF', 'CONF:CAP 1 UF', 'C334', 'D334', 'DET:BAND 20', '5', '0.5')
        thread = message('Переключите провода\n для измерения тока')
        thread = reset()
        # 0.001A
        thread = call('dci', '0.0001', 'OUT 0.0001 A', 'CONF:CURR:DC 0.001', 'C176', 'D176', 'DET:BAND 20', '5', '0.1')
        thread = call('dci', '0.0003', 'OUT 0.0003 A', 'CONF:CURR:DC 0.001', 'C177', 'D177', 'DET:BAND 20', '5', '0.067')
        thread = call('dci', '0.0005', 'OUT 0.0005 A', 'CONF:CURR:DC 0.001', 'C178', 'D178', 'DET:BAND 20', '5', '0.06')
        thread = call('dci', '0.0007', 'OUT 0.0007 A', 'CONF:CURR:DC 0.001', 'C179', 'D179', 'DET:BAND 20', '5', '0.057')
        thread = call('dci', '0.001', 'OUT 0.001 A', 'CONF:CURR:DC 0.001', 'C180', 'D180', 'DET:BAND 20', '5', '0.055')
        # 0.01A
        thread = call('dci', '0.001', 'OUT 0.001 A', 'CONF:CURR:DC 0.01', 'C181', 'D181', 'DET:BAND 20', '5', '0.25')
        thread = call('dci', '0.003', 'OUT 0.003 A', 'CONF:CURR:DC 0.01', 'C182', 'D182', 'DET:BAND 20', '5', '0.117')
        thread = call('dci', '0.005', 'OUT 0.005 A', 'CONF:CURR:DC 0.01', 'C183', 'D183', 'DET:BAND 20', '5', '0.09')
        thread = call('dci', '0.007', 'OUT 0.007 A', 'CONF:CURR:DC 0.01', 'C184', 'D184', 'DET:BAND 20', '5', '0.079')
        thread = call('dci', '0.01', 'OUT 0.01 A', 'CONF:CURR:DC 0.01', 'C185', 'D185', 'DET:BAND 20', '5', '0.07')
        # 0.1A
        thread = call('dci', '0.01', 'OUT 0.01 A', 'CONF:CURR:DC 0.1', 'C186', 'D186', 'DET:BAND 20', '5', '0.1')
        thread = call('dci', '0.03', 'OUT 0.03 A', 'CONF:CURR:DC 0.1', 'C187', 'D187', 'DET:BAND 20', '5', '0.067')
        thread = call('dci', '0.05', 'OUT 0.05 A', 'CONF:CURR:DC 0.1', 'C188', 'D188', 'DET:BAND 20', '5', '0.06')
        thread = call('dci', '0.07', 'OUT 0.07 A', 'CONF:CURR:DC 0.1', 'C189', 'D189', 'DET:BAND 20', '5', '0.057')
        thread = call('dci', '0.1', 'OUT 0.1 A', 'CONF:CURR:DC 0.1', 'C190', 'D190', 'DET:BAND 20', '5', '0.055')
        # 1A
        thread = call('dci', '0.1', 'OUT 0.1 A', 'CONF:CURR:DC 1.0', 'C191', 'D191', 'DET:BAND 20', '5', '0.18')
        thread = call('dci', '0.3', 'OUT 0.3 A', 'CONF:CURR:DC 1.0', 'C192', 'D192', 'DET:BAND 20', '5', '0.113')
        thread = call('dci', '0.5', 'OUT 0.5 A', 'CONF:CURR:DC 1.0', 'C193', 'D193', 'DET:BAND 20', '5', '0.1')
        thread = call('dci', '0.7', 'OUT 0.7 A', 'CONF:CURR:DC 1.0', 'C194', 'D194', 'DET:BAND 20', '5', '0.094')
        thread = call('dci', '1.0', 'OUT 1.0 A', 'CONF:CURR:DC 1.0', 'C195', 'D195', 'DET:BAND 20', '5', '0.09')
        # 3A
        thread = call('dci', '0.3', 'OUT 0.3 A', 'CONF:CURR:DC 3.0', 'C196', 'D196', 'DET:BAND 20', '5', '0.4')
        thread = call('dci', '0.9', 'OUT 0.9 A', 'CONF:CURR:DC 3.0', 'C197', 'D197', 'DET:BAND 20', '5', '0.267')
        thread = call('dci', '1.5', 'OUT 1.5 A', 'CONF:CURR:DC 3.0', 'C198', 'D198', 'DET:BAND 20', '5', '0.24')
        thread = call('dci', '2.1', 'OUT 2.1 A', 'CONF:CURR:DC 3.0', 'C199', 'D199', 'DET:BAND 20', '5', '0.229')
        if b1[1] == '5500E':
            thread = call('dci', '2.85', 'OUT 2.85 A', 'CONF:CURR:DC 3.0', 'C200', 'D200', 'DET:BAND 20', '5', '0.232')
        thread = reset()
        # ~0.0001A, 20Hz
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 20 Hz', 'CONF:CURR:AC 0.0001', 'D210', 'E210', 'DET:BAND 3', '8', '0.14')
        # ~0.001A, 20Hz
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 20 Hz', 'CONF:CURR:AC 0.001', 'D211', 'E211', 'DET:BAND 3', '8', '0.5')
        thread = call('aci', '0.0003', 'OUT 0.0003 A, 20 Hz', 'CONF:CURR:AC 0.001', 'D212', 'E212', 'DET:BAND 3', '8', '0.233')
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 20 Hz', 'CONF:CURR:AC 0.001', 'D213', 'E213', 'DET:BAND 3', '8', '0.18')
        thread = call('aci', '0.0007', 'OUT 0.0007 A, 20 Hz', 'CONF:CURR:AC 0.001', 'D214', 'E214', 'DET:BAND 3', '8', '0.157')
        thread = call('aci', '0.001', 'OUT 0.001 A, 20 Hz', 'CONF:CURR:AC 0.001', 'D215', 'E215', 'DET:BAND 3', '8', '0.14')
        # ~0.01A, 20Hz
        thread = call('aci', '0.001', 'OUT 0.001 A, 20 Hz', 'CONF:CURR:AC 0.01', 'D216', 'E216', 'DET:BAND 3', '8', '0.5')
        thread = call('aci', '0.003', 'OUT 0.003 A, 20 Hz', 'CONF:CURR:AC 0.01', 'D217', 'E217', 'DET:BAND 3', '8', '0.233')
        thread = call('aci', '0.005', 'OUT 0.005 A, 20 Hz', 'CONF:CURR:AC 0.01', 'D218', 'E218', 'DET:BAND 3', '8', '0.18')
        thread = call('aci', '0.007', 'OUT 0.007 A, 20 Hz', 'CONF:CURR:AC 0.01', 'D219', 'E219', 'DET:BAND 3', '8', '0.157')
        thread = call('aci', '0.01', 'OUT 0.01 A, 20 Hz', 'CONF:CURR:AC 0.01', 'D220', 'E220', 'DET:BAND 3', '8', '0.14')
        # ~0.1A, 20Hz
        thread = call('aci', '0.01', 'OUT 0.01 A, 20 Hz', 'CONF:CURR:AC 0.1', 'D221', 'E221', 'DET:BAND 3', '8', '0.5')
        thread = call('aci', '0.03', 'OUT 0.03 A, 20 Hz', 'CONF:CURR:AC 0.1', 'D222', 'E222', 'DET:BAND 3', '8', '0.233')
        thread = call('aci', '0.05', 'OUT 0.05 A, 20 Hz', 'CONF:CURR:AC 0.1', 'D223', 'E223', 'DET:BAND 3', '8', '0.18')
        thread = call('aci', '0.07', 'OUT 0.07 A, 20 Hz', 'CONF:CURR:AC 0.1', 'D224', 'E224', 'DET:BAND 3', '8', '0.157')
        thread = call('aci', '0.1', 'OUT 0.1 A, 20 Hz', 'CONF:CURR:AC 0.1', 'D225', 'E225', 'DET:BAND 3', '8', '0.14')
        # ~1A, 20Hz
        thread = call('aci', '0.1', 'OUT 0.1 A, 20 Hz', 'CONF:CURR:AC 1.0', 'D226', 'E226', 'DET:BAND 3', '8', '0.5')
        thread = call('aci', '0.3', 'OUT 0.3 A, 20 Hz', 'CONF:CURR:AC 1.0', 'D227', 'E227', 'DET:BAND 3', '8', '0.233')
        thread = call('aci', '0.5', 'OUT 0.5 A, 20 Hz', 'CONF:CURR:AC 1.0', 'D228', 'E228', 'DET:BAND 3', '8', '0.18')
        thread = call('aci', '0.7', 'OUT 0.7 A, 20 Hz', 'CONF:CURR:AC 1.0', 'D229', 'E229', 'DET:BAND 3', '8', '0.157')
        thread = call('aci', '1.0', 'OUT 1.0 A, 20 Hz', 'CONF:CURR:AC 1.0', 'D230', 'E230', 'DET:BAND 3', '8', '0.14')
        # ~0.0001A, 1kHz
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 1 kHz', 'CONF:CURR:AC 0.0001', 'D231', 'E231', 'DET:BAND 20', '5', '0.14')
        # ~0.001A, 1kHz
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 1 kHz', 'CONF:CURR:AC 0.001', 'D232', 'E232', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.0003', 'OUT 0.0003 A, 1 kHz', 'CONF:CURR:AC 0.001', 'D233', 'E233', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 1 kHz', 'CONF:CURR:AC 0.001', 'D234', 'E234', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.0007', 'OUT 0.0007 A, 1 kHz', 'CONF:CURR:AC 0.001', 'D235', 'E235', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '0.001', 'OUT 0.001 A, 1 kHz', 'CONF:CURR:AC 0.001', 'D236', 'E236', 'DET:BAND 20', '5', '0.14')
        # ~0.01A, 1kHz
        thread = call('aci', '0.001', 'OUT 0.001 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D237', 'E237', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.003', 'OUT 0.003 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D238', 'E238', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.005', 'OUT 0.005 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D239', 'E239', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.007', 'OUT 0.007 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D240', 'E240', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '0.01', 'OUT 0.01 A, 1 kHz', 'CONF:CURR:AC 0.01', 'D241', 'E241', 'DET:BAND 20', '5', '0.14')
        # ~0.1A, 1kHz
        thread = call('aci', '0.01', 'OUT 0.01 A, 1 kHz', 'CONF:CURR:AC 0.1', 'D242', 'E242', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.03', 'OUT 0.03 A, 1 kHz', 'CONF:CURR:AC 0.1', 'D243', 'E243', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.05', 'OUT 0.05 A, 1 kHz', 'CONF:CURR:AC 0.1', 'D244', 'E244', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.07', 'OUT 0.07 A, 1 kHz', 'CONF:CURR:AC 0.1', 'D245', 'E245', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '0.1', 'OUT 0.1 A, 1 kHz', 'CONF:CURR:AC 0.1', 'D246', 'E246', 'DET:BAND 20', '5', '0.14')
        # ~1A, 1kHz
        thread = call('aci', '0.1', 'OUT 0.1 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D247', 'E247', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.3', 'OUT 0.3 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D248', 'E248', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.5', 'OUT 0.5 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D249', 'E249', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.7', 'OUT 0.7 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D250', 'E250', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '1.0', 'OUT 1.0 A, 1 kHz', 'CONF:CURR:AC 1.0', 'D251', 'E251', 'DET:BAND 20', '5', '0.14')
        # ~3A, 1kHz
        thread = call('aci', '0.3', 'OUT 0.3 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D252', 'E252', 'DET:BAND 20', '5', '0.63')
        thread = call('aci', '0.9', 'OUT 0.9 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D253', 'E253', 'DET:BAND 20', '5', '0.363')
        thread = call('aci', '1.5', 'OUT 1.5 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D254', 'E254', 'DET:BAND 20', '5', '0.31')
        thread = call('aci', '2.1', 'OUT 2.1 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D255', 'E255', 'DET:BAND 20', '5', '0.287')
        if b1[1] == '5500E':
            thread = call('aci', '2.85', 'OUT 2.85 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D256', 'E256', 'DET:BAND 20', '5', '0.284')
        # ~0.0001A, 5kHz
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 5 kHz', 'CONF:CURR:AC 0.0001', 'D262', 'E262', 'DET:BAND 20', '5', '0.14')
        # ~0.001A, 5kHz
        thread = call('aci', '0.0001', 'OUT 0.0001 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D263', 'E263', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.0003', 'OUT 0.0003 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D264', 'E264', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.0005', 'OUT 0.0005 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D265', 'E265', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.0007', 'OUT 0.0007 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D266', 'E266', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '0.001', 'OUT 0.001 A, 5 kHz', 'CONF:CURR:AC 0.001', 'D267', 'E267', 'DET:BAND 20', '5', '0.14')
        # ~0.01A, 5kHz
        thread = call('aci', '0.001', 'OUT 0.001 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D268', 'E268', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.003', 'OUT 0.003 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D269', 'E269', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.005', 'OUT 0.005 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D270', 'E270', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.007', 'OUT 0.007 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D271', 'E271', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '0.01', 'OUT 0.01 A, 5 kHz', 'CONF:CURR:AC 0.01', 'D272', 'E272', 'DET:BAND 20', '5', '0.14')
        # ~0.1A, 5kHz
        thread = call('aci', '0.01', 'OUT 0.01 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D273', 'E273', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.03', 'OUT 0.03 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D274', 'E274', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.05', 'OUT 0.05 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D275', 'E275', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.07', 'OUT 0.07 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D276', 'E276', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '0.1', 'OUT 0.1 A, 5 kHz', 'CONF:CURR:AC 0.1', 'D277', 'E277', 'DET:BAND 20', '5', '0.14')
        # ~1A, 5kHz
        thread = call('aci', '0.1', 'OUT 0.1 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D278', 'E278', 'DET:BAND 20', '5', '0.5')
        thread = call('aci', '0.3', 'OUT 0.3 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D279', 'E279', 'DET:BAND 20', '5', '0.233')
        thread = call('aci', '0.5', 'OUT 0.5 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D280', 'E280', 'DET:BAND 20', '5', '0.18')
        thread = call('aci', '0.7', 'OUT 0.7 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D281', 'E281', 'DET:BAND 20', '5', '0.157')
        thread = call('aci', '1.0', 'OUT 1.0 A, 5 kHz', 'CONF:CURR:AC 1.0', 'D282', 'E282', 'DET:BAND 20', '5', '0.14')
        # ~3A, 5kHz
        thread = call('aci', '0.3', 'OUT 0.3 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D283', 'E283', 'DET:BAND 20', '5', '0.63')
        thread = call('aci', '0.9', 'OUT 0.9 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D284', 'E284', 'DET:BAND 20', '5', '0.363')
        thread = call('aci', '1.5', 'OUT 1.5 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D285', 'E285', 'DET:BAND 20', '5', '0.31')
        thread = call('aci', '2.1', 'OUT 2.1 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D286', 'E286', 'DET:BAND 20', '5', '0.287')
        if b1[1] == '5522A':
            thread = message('Переключите красный провод на калибраторе в разъем больше 2,5 А')
            thread = call('dci', '2.85', 'OUT 2.85 A', 'CONF:CURR:DC 3.0', 'C200', 'D200', 'DET:BAND 20', '5', '0.232')
            thread = call('aci', '2.85', 'OUT 2.85 A, 1 kHz', 'CONF:CURR:AC 3.0', 'D256', 'E256', 'DET:BAND 20', '5', '0.284')
            thread = call('aci', '2.85', 'OUT 2.85 A, 5 kHz', 'CONF:CURR:AC 3.0', 'D287', 'E287', 'DET:BAND 20', '5', '0.284')
        # 10A
        if b1[1] == '5500E':
            thread = message('Переключите красный провод на контакт 10А мультиметра')
            thread = reset()
            thread = call('dci', '1.0', 'OUT 1.0 A', 'CONF:CURR:DC 10.0', 'C201', 'D201', 'DET:BAND 20', '5', '0.22')
            thread = call('dci', '3.0', 'OUT 3.0 A', 'CONF:CURR:DC 10.0', 'C202', 'D202', 'DET:BAND 20', '5', '0.153')
            thread = call('dci', '5.0', 'OUT 5.0 A', 'CONF:CURR:DC 10.0', 'C203', 'D203', 'DET:BAND 20', '5', '0.14')
            thread = call('dci', '7.0', 'OUT 7.0 A', 'CONF:CURR:DC 10.0', 'C204', 'D204', 'DET:BAND 20', '5', '0.191')
            thread = call('dci', '10.0', 'OUT 10.0 A', 'CONF:CURR:DC 10.0', 'C205', 'D205', 'DET:BAND 20', '5', '0.23')
        # ~10A, 1kHz
            thread = call('aci', '1', 'OUT 1 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D257', 'E257', 'DET:BAND 20', '5', '0.5')
            thread = call('aci', '3', 'OUT 3 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D258', 'E258', 'DET:BAND 20', '5', '0.233')
            thread = call('aci', '5', 'OUT 5 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D259', 'E259', 'DET:BAND 20', '5', '0.18')
            thread = call('aci', '7', 'OUT 7 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D260', 'E260', 'DET:BAND 20', '5', '0.214')
            thread = call('aci', '10', 'OUT 10 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D261', 'E261', 'DET:BAND 20', '5', '0.24')
        elif b1[1] == '5522A':
            thread = message('Переключите красный провод на контакт 10А мультиметра')
            thread = reset()
            thread = call('dci', '1.0', 'OUT 1.0 A', 'CONF:CURR:DC 10.0', 'C201', 'D201', 'DET:BAND 20', '5', '0.22')
            thread = call('aci', '1', 'OUT 1 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D257', 'E257', 'DET:BAND 20', '5', '0.5')
            thread = call('aci', '1', 'OUT 1 A, 5 kHz', 'CONF:CURR:AC 10.0', 'D288', 'E288', 'DET:BAND 20', '5', '0.5')
            thread = message('Переключите красный провод на калибраторе в разъем больше 2,5 А')
            thread = call('dci', '3.0', 'OUT 3.0 A', 'CONF:CURR:DC 10.0', 'C202', 'D202', 'DET:BAND 20', '5', '0.153')
            thread = call('dci', '5.0', 'OUT 5.0 A', 'CONF:CURR:DC 10.0', 'C203', 'D203', 'DET:BAND 20', '5', '0.14')
            thread = call('dci', '7.0', 'OUT 7.0 A', 'CONF:CURR:DC 10.0', 'C204', 'D204', 'DET:BAND 20', '5', '0.191')
            thread = call('dci', '10.0', 'OUT 10.0 A', 'CONF:CURR:DC 10.0', 'C205', 'D205', 'DET:BAND 20', '5', '0.23')
            thread = call('aci', '3', 'OUT 3 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D258', 'E258', 'DET:BAND 20', '5', '0.233')
            thread = call('aci', '5', 'OUT 5 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D259', 'E259', 'DET:BAND 20', '5', '0.18')
            thread = call('aci', '7', 'OUT 7 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D260', 'E260', 'DET:BAND 20', '5', '0.214')
            thread = call('aci', '10', 'OUT 10 A, 1 kHz', 'CONF:CURR:AC 10.0', 'D261', 'E261', 'DET:BAND 20', '5', '0.24')
            thread = call('aci', '3', 'OUT 3 A, 5 kHz', 'CONF:CURR:AC 10.0', 'D289', 'E289', 'DET:BAND 20', '5', '0.233')
            thread = call('aci', '5', 'OUT 5 A, 5 kHz', 'CONF:CURR:AC 10.0', 'D290', 'E290', 'DET:BAND 20', '5', '0.18')
            thread = call('aci', '7', 'OUT 7 A, 5 kHz', 'CONF:CURR:AC 10.0', 'D291', 'E291', 'DET:BAND 20', '5', '0.214')
            thread = call('aci', '10', 'OUT 10 A, 5 kHz', 'CONF:CURR:AC 10.0', 'D292', 'E292', 'DET:BAND 20', '5', '0.24')
        thread = message('Переключите провода по четырехпроводной схеме\n для измерения сопротивления')
        thread = reset()	
        # Ohm
        thread = call('res4', '100', 'OUT 100 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'C311', 'D311', 'DET:BAND 20', '5', '0.01')
        thread = call('res4', '1000', 'OUT 1 KOHM; ZCOMP WIRE4', 'CONF:FRES 1 KOHM', 'C312', 'D312', 'DET:BAND 20', '5', '0.005')
        thread = call('res4', '10000', 'OUT 10 KOHM; ZCOMP WIRE4', 'CONF:FRES 10 KOHM', 'C313', 'D313', 'DET:BAND 20', '5', '0.005')
        thread = call('res4', '100000', 'OUT 100 KOHM; ZCOMP WIRE4', 'CONF:FRES 100 KOHM', 'C314', 'D314', 'DET:BAND 20', '5', '0.005')
        thread = message('Переключите провода по двухпроводной схеме\n для измерения сопротивления')
        thread = reset()
        thread = call('res2', '1000000', 'OUT 1 MOHM; ZCOMP WIRE2', 'CONF:RES 1 MOHM', 'C322', 'D322', 'DET:BAND 20', '5', '0.075')
        thread = call('res2', '10000000', 'OUT 10 MOHM; ZCOMP WIRE2', 'CONF:RES 10 MOHM', 'C323', 'D323', 'DET:BAND 20', '5', '0.026')
        thread = call('res2', '100000000', 'OUT 100 MOHM; ZCOMP WIRE2', 'CONF:RES 100 MOHM', 'C324', 'D324', 'DET:BAND 20', '5', '0.301')
        if b1[1] == '5522A':
            thread = call('res2', '1000000000', 'OUT 1 GOHM; ZCOMP WIRE2', 'CONF:RES 1 GOHM', 'C325', 'D325', 'DET:BAND 20', '5', '3.001')
        thread = message('Калибровка завершена')
        thread = reset()


def start_thread(fun, a=(), k={}):
    threading.Thread(target=fun, args=a, kwargs=k).start()


def tkloop():
    try:
        while True:
            f, a, k = q.get_nowait()
            f(*a, **k)
    except:
        pass

    root.after(100, tkloop)

def protokol(): 
    rep = filedialog.askopenfilenames(
        parent=root,
        initialdir='C:\ITL\DMM\Protocol',
        initialfile='',
        filetypes=[("xlsx", "*.xlsx"),("All files", "*")])
    try:
	    os.startfile(rep[0])
    except IndexError:
        print("No file selected")

def about_win():
    top = Toplevel(root)
    top.title('О программе')
    top.iconbitmap('icon/icon.ico')
    top.resizable(0, 0)
    w = top.winfo_screenwidth()
    h = top.winfo_screenheight()
    w = w // 3
    h = h // 2
    w = w - 200
    h = h - 200
    top.geometry('270x225+{}+{}'.format(w, h))

    text1 = ('Digital Multimeter v1.07\rAutor: ITL\r\rПоддерживаемые мультиметры:\r34401A\r34410A\r34411A\r34460A\r34461A\r34465A\r34470A')

    mick = ttk.Label(top, image=img5).place(x=10,y=10)
    autor = ttk.Label(top, justify=LEFT, text=text1).place(x=60,y=5)
    But = ttk.Button(top, text='OK', width=10, command=top.destroy).place(x=80,y=180)  

    top.transient(root)
    top.grab_set()
    root.wait_window(top)


class AnimatedGif(object):
    def __init__(self, image_file_path):
        self._frames = []

        frame_num = 0
        while True:
            try:
                frame = PhotoImage(file=image_file_path,
                                   format='gif -index {}'.format(frame_num))
            except TclError:
                break
            self._frames.append(frame)
            frame_num += 1

    def __len__(self):
        return len(self._frames)

    def __getitem__(self, frame_num):
        return self._frames[frame_num]


def update_label_image(label, img4, ms_delay, frame_num):
    global cancel_id
    label.configure(image=img4[frame_num])
    frame_num = (frame_num+1) % len(img4)
    cancel_id = root.after(
        ms_delay, update_label_image, label, img4, ms_delay, frame_num)

def enable_animation():
    global cancel_id
    if cancel_id is None:
        ms_delay = 1000 // len(img4)
        cancel_id = root.after(
            ms_delay, update_label_image, animation, img4, ms_delay, 0)

def cancel_animation():
    global cancel_id
    if cancel_id is not None:
        root.after_cancel(cancel_id)
        cancel_id = None


root = Tk()
root.title('Digital Multimeter')
root.geometry('805x430')
root.iconbitmap('icon/icon.ico')
root.configure(background='FloralWhite')
root.resizable(width=False, height=False)
frame = Frame(root)
frame.grid()

#ttk.Style().theme_use('alt')
ttk.Style().configure('TButton', padding=6, font='arial 10', foreground='black', background='FloralWhite')
ttk.Style().configure('BW.Label', padding=6, font='arial 10', foreground='black', background='FloralWhite')
ttk.Style().configure('TLabelframe', background='FloralWhite')
ttk.Style().configure("TProgressbar", foreground='blue', background='blue')

main_menu = Menu(root)

file_menu = Menu(main_menu, tearoff=False)
file_menu.add_command(label='Новый')
file_menu.add_command(label='Открыть')
file_menu.add_command(label='Сохранить')
file_menu.add_separator()
file_menu.add_command(label='Закрыть', command=root.quit)

main_menu.add_cascade(label='Файл', menu=file_menu)
main_menu.add_cascade(label='Протокол', command=protokol)
main_menu.add_cascade(label='Настройки')
main_menu.add_cascade(label='О программе', command=about_win)

today = datetime.datetime.today()
a = StringVar()
b = StringVar()
# c = StringVar()
d = today.strftime('%d.%m.%Y,%H.%M.%S')
d1 = today.strftime('%H:%M:%S')
e = today.strftime('%d.%m.%Y')
f = StringVar()
g = StringVar()
h = StringVar()
k = StringVar()
l = StringVar()
m = StringVar()
n = StringVar()
a10 = StringVar()
b10 = StringVar()
b14 = StringVar()
b15 = StringVar()

img1 = PhotoImage(file='icon/pan2.gif')
img2 = PhotoImage(file='icon/start1.gif')
img3 = PhotoImage(file='icon/ref1.gif')
img4 = AnimatedGif('icon/progress.gif')
img5 = PhotoImage(file='icon/mick.gif')
cancel_id = None
label = Label(root, image=img1)
label.place(x=1, y=1)

lbf1 = ttk.LabelFrame(root, text='Идентификация', width=475, height=185, style='TLabelframe')
lbf1.place(x=5, y=58)
lbf2 = ttk.LabelFrame(root, text='Условия калибровки', width=475, height=80, style='TLabelframe')
lbf2.place(x=5, y=250)
lbf3 = ttk.LabelFrame(root, text='Прогресс', width=645, height=50, style='TLabelframe')
lbf3.place(x=5, y=355)

but1 = ttk.Button(root, text='Connect DMM', width=16, command=connect_dmm, style='TButton')
but1.place(x=10, y=75)
but2 = ttk.Button(root, text='Connect Fluke', width=16, command=connect_fluke, style='TButton')
but2.place(x=10, y=115)
but3 = Button(root, image=img2, command=start, style='TButton')
but3.place(x=744, y=362)
but4 = Button(root, image=img3, command=pribor, style='TButton')
but4.place(x=695, y=362)

lab1 = ttk.Label(root, text='ID DMM:', style='BW.Label')
lab1.place(x=47, y=157)
lab2 = ttk.Label(root, text='ID Fluke:', style='BW.Label')
lab2.place(x=47, y=180)
lab3 = ttk.Label(root, text='Название протокола:', style='BW.Label')
lab3.place(x=7, y=210)
lab4 = ttk.Label(root, text='Температура:', style='BW.Label')
lab4.place(x=7, y=265)
lab5 = ttk.Label(root, text='Влажность:', style='BW.Label')
lab5.place(x=100, y=265)
lab6 = ttk.Label(root, text='Давление:', style='BW.Label')
lab6.place(x=190, y=265)
lab7 = ttk.Label(root, text='Заказчик:', style='BW.Label')
lab7.place(x=275, y=265)
lab8 = ttk.Label(root, text='Поверитель:', style='BW.Label')
lab8.place(x=360, y=265)
animation = ttk.Label(root,image=img4[0], style='TLabel')
animation.place(x=665, y=370)

combo1 = ttk.Combobox(root, state='readonly', height=4, width=50)
combo1.place(x=145, y=82)
combo2 = ttk.Combobox(root, textvariable=g, state='readonly', height=4, width=50)
combo2.place(x=145, y=122)

entry1 = ttk.Entry(root, textvariable=a, state='readonly', width=53, font='arial 8')
entry1.place(x=145, y=155)
entry2 = ttk.Entry(root, textvariable=b, state='readonly', width=53, font='arial 8')
entry2.place(x=145, y=182)
entry3 = ttk.Entry(root, textvariable=f, width=53, font='arial 8')
entry3.place(x=145, y=212)
entry4 = ttk.Entry(root, textvariable=h, width=10, font='arial 8')
entry4.place(x=15, y=295)
entry5 = ttk.Entry(root, textvariable=k, width=10, font='arial 8')
entry5.place(x=110, y=295)
entry6 = ttk.Entry(root, textvariable=l, width=10, font='arial 8')
entry6.place(x=200, y=295)
entry7 = ttk.Entry(root, textvariable=m, width=10, font='arial 8')
entry7.place(x=285, y=295)
entry8 = ttk.Entry(root, textvariable=n, width=17, font='arial 8')
entry8.place(x=365, y=295)

progress1 = ttk.Progressbar(root, orient='horizontal', mode='determinate', length = 640, value = 0, style='TProgressbar')
progress1.place(x=7, y=375)

lb = Listbox(root, selectmode=EXTENDED, width=49, height=18, relief=RIDGE)
lb.place(x=490, y=66)

root.event_add('<<Paste>>', '<Control-igrave>')
root.event_add("<<Copy>>", "<Control-ntilde>")
root.config(menu=main_menu)
tkloop()
pribor()

root.mainloop()
